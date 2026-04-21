"""
=============================================================================
SymfoniaYearEndAuditor – Automatyczna Kontrola Jakości Danych Finansowych
Biuro Rachunkowe Abacus | Zamknięcie Roku Obrachunkowego
=============================================================================
Wersja 3.3 – uniwersalny parser wyciągów: obsługa zestawień operacji (mBank)
             obok klasycznych wyciągów. Obsługa tekstu bez spacji w PDF.

Wersja 3.2 – trzy krytyczne kontrole zamknięcia roku:

  R1: Weryfikacja konta 860 (Wynik finansowy) z auto-detekcją stanu
      - Przed zamknięciem: saldo 860 = L z RZiS
      - Po zamknięciu: obroty Ma − Wn = L z RZiS
      - Ostrzeżenie gdy 860 puste a grupy 4/7 mają salda

  R2: Kontrola dwustronnego zapisu (Wn = Ma) dla wszystkich 3 par kolumn
      - Bilans Otwarcia (BO)
      - Obroty bieżącego miesiąca
      - Obroty narastająco

  R3: Weryfikacja kapitału zakładowego vs oficjalne API KRS
      - Integracja z api-krs.ms.gov.pl (Ministerstwo Sprawiedliwości)
      - Rozpoznawanie konta "kapitał zakładowy/podstawowy" po opisie
      - Fallback na pozycję A Bilansu pasywów gdy brak konta w ZOiS

Wcześniejsze funkcje (v3.1):
  - Parser ZOiS/Bilans/RZiS z PDF Symfonii (z dekodowaniem CID polskich znaków)
  - Rozpoznawanie kont po opisie (odbiorcy, dostawcy, ZUS, PIT, wynagrodzenia)
  - Reguły krzyżowe: konto 700 ≈ RZiS A, grupa 4 ≈ RZiS B, RZiS L = Bilans A.VI
  - Weryfikacja wielu rachunków bankowych (analityki 130-X)
=============================================================================
"""

from __future__ import annotations

import io
import re
import logging
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation
from enum import Enum
from typing import Dict, List, Optional, Tuple, Union

import pandas as pd

# ── Opcjonalne biblioteki ─────────────────────────────────────────────────────
try:
    import pdfplumber
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

logger = logging.getLogger("SymfoniaAuditor")
logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(message)s")

MIESIACE_PL = {
    1: "styczeń", 2: "luty", 3: "marzec", 4: "kwiecień",
    5: "maj", 6: "czerwiec", 7: "lipiec", 8: "sierpień",
    9: "wrzesień", 10: "październik", 11: "listopad", 12: "grudzień",
}

# Regex dla kwoty w formacie polskim: 1.114.621,54 | 0,00 | -112.140,00
RE_KWOTA = r"-?\d+(?:\.\d{3})*(?:,\d{1,2})?"


# =============================================================================
# DEKODER POLSKICH ZNAKÓW Z PDF SYMFONII
# =============================================================================
# Symfonia eksportuje PDF z własną czcionką gdzie polskie znaki są kodowane
# jako glyph-id'y. pdfplumber zwraca je jako "(cid:XXX)". Ta mapa przekłada
# najczęściej występujące glyphy na prawdziwe polskie litery.
CID_TO_PL = {
    # Niestandardowe CID dla zwykłych liter łacińskich (anomalia fontu Symfonii)
    39:  "D",   71:  "D",   79:  "l",   88:  "u",
    # Polskie znaki – zidentyfikowane z plików Abacus i IGRAPES
    211: "Ó",   243: "ó",
    224: "Ł",   225: "ł",
    260: "Ą",   261: "ą",
    262: "Ć",   252: "ć",   253: "Ć",
    265: "Ę",   266: "ę",   267: "Ę",
    269: "Ń",   276: "ń",   277: "Ń",
    285: "Ś",   286: "ś",
    297: "Ź",   240: "ź",   241: "Ź",
    298: "ż",   300: "Ż",
}
_RE_CID = re.compile(r"\(cid:(\d+)\)")


def dekoduj_cid(tekst: str) -> str:
    """
    Zamienia glyph-id Symfonii na polskie znaki (np. '(cid:266)' → 'ę').
    Nieznane CID zostają zastąpione znakiem zapytania.
    """
    if not tekst or "(cid:" not in tekst:
        return tekst
    return _RE_CID.sub(lambda m: CID_TO_PL.get(int(m.group(1)), "?"), tekst)


# =============================================================================
# TYPY
# =============================================================================

class StatusAudytu(Enum):
    OK       = "✅ OK"
    BLAD     = "❌ BŁĄD"
    OSTRZEZ  = "⚠️  OSTRZEŻENIE"
    BRAK     = "🔍 BRAK DANYCH"
    INFO     = "ℹ️  INFO"


@dataclass
class PunktKontroli:
    konto:   str
    punkt:   str
    status:  StatusAudytu
    uwagi:   str = ""
    wartosc: str = ""


@dataclass
class SumyRazemZOiS:
    """
    Wiersz "Suma razem" z końcowej strony ZOiS (Symfonia).
    Zawiera oficjalne sumy dla wszystkich 4 par kolumn – używane do weryfikacji
    dwustronnego zapisu (Wn = Ma musi być spełnione dla BO, obrotów miesiąca
    i obrotów narastająco).
    """
    bo_wn:           Decimal = Decimal("0")
    bo_ma:           Decimal = Decimal("0")
    obroty_wn:       Decimal = Decimal("0")
    obroty_ma:       Decimal = Decimal("0")
    narastajaco_wn:  Decimal = Decimal("0")
    narastajaco_ma:  Decimal = Decimal("0")
    saldo_wn:        Decimal = Decimal("0")
    saldo_ma:        Decimal = Decimal("0")
    wykryto:         bool = False   # Czy linia "Suma razem" została znaleziona


@dataclass
class DaneKRS:
    """
    Dane pobrane z oficjalnego API KRS (Ministerstwo Sprawiedliwości).
    Endpoint: https://api-krs.ms.gov.pl/api/krs/OdpisAktualny/{krs}?rejestr=P&format=json
    """
    numer_krs:           str = ""
    nazwa:               str = ""
    forma_prawna:        str = ""
    nip:                 str = ""
    regon:               str = ""
    kapital_zakladowy:   Decimal = Decimal("0")
    data_pobrania:       str = ""
    blad:                Optional[str] = None   # Komunikat błędu jeśli pobranie się nie udało


@dataclass
class DaneZOiS:
    """
    Zestawienie Obrotów i Sald z Symfonii – syntetyki + analityki.
    """
    konta: Dict[str, Tuple[Decimal, Decimal]] = field(default_factory=dict)
    konta_analityki: Dict[str, Tuple[Decimal, Decimal]] = field(default_factory=dict)
    opisy: Dict[str, str] = field(default_factory=dict)
    sumy_razem: SumyRazemZOiS = field(default_factory=SumyRazemZOiS)
    # Obroty narastająco per syntetyka – używane do weryfikacji przeksięgowań 860
    obroty_narastajaco: Dict[str, Tuple[Decimal, Decimal]] = field(default_factory=dict)

    def pobierz_konta_bankowe(self) -> List[Tuple[str, str, Decimal]]:
        """Zwraca listę wszystkich rachunków bankowych (analityki 130)."""
        wynik = []
        for numer, (wn, ma) in self.konta_analityki.items():
            if normalize_konto(numer) == "130":
                saldo_netto = wn - ma
                opis = self.opisy.get(numer, "Rachunek bankowy")
                wynik.append((numer, opis, saldo_netto))
        if not wynik and "130" in self.konta:
            wn, ma = self.konta["130"]
            wynik.append(("130", self.opisy.get("130", "Rachunek bankowy"), wn - ma))
        return sorted(wynik, key=lambda x: x[0])


@dataclass
class DaneBilansu:
    """
    Dane Bilansu sparsowane z Symfonii.
    Zawiera zarówno szczegółowe pozycje (A.aktywa trwałe, B.obrotowe, C, D)
    jak i sumy bilansowe. Wynik netto pobierany z pasywów A.VI.
    """
    # AKTYWA (sekcja AKTYWA w PDF)
    aktywa_trwale_biezacy:    Decimal = Decimal("0")   # poz. A
    aktywa_trwale_ubiegly:    Decimal = Decimal("0")
    aktywa_obrotowe_biezacy:  Decimal = Decimal("0")   # poz. B
    aktywa_obrotowe_ubiegly:  Decimal = Decimal("0")
    nalezne_wplaty_biezacy:   Decimal = Decimal("0")   # poz. C
    nalezne_wplaty_ubiegly:   Decimal = Decimal("0")
    udzialy_wlasne_biezacy:   Decimal = Decimal("0")   # poz. D
    udzialy_wlasne_ubiegly:   Decimal = Decimal("0")
    suma_aktywow_biezacy:     Decimal = Decimal("0")
    suma_aktywow_ubiegly:     Decimal = Decimal("0")

    # PASYWA (sekcja PASYWA w PDF)
    kapital_wlasny_biezacy:   Decimal = Decimal("0")   # poz. A
    kapital_wlasny_ubiegly:   Decimal = Decimal("0")
    zobowiazania_biezacy:     Decimal = Decimal("0")   # poz. B
    zobowiazania_ubiegly:     Decimal = Decimal("0")
    suma_pasywow_biezacy:     Decimal = Decimal("0")
    suma_pasywow_ubiegly:     Decimal = Decimal("0")

    # Wynik netto z Bilansu (pasywa A.VI) – do weryfikacji krzyżowej z RZiS
    wynik_netto_biezacy:      Decimal = Decimal("0")
    wynik_netto_ubiegly:      Decimal = Decimal("0")

    # ── Kompatybilność wsteczna z v2.0 ───────────────────────────────────────
    @property
    def aktywa_biezacy(self) -> Decimal: return self.suma_aktywow_biezacy
    @property
    def pasywa_biezacy(self) -> Decimal: return self.suma_pasywow_biezacy
    @property
    def aktywa_ubiegly(self) -> Decimal: return self.suma_aktywow_ubiegly
    @property
    def pasywa_ubiegly(self) -> Decimal: return self.suma_pasywow_ubiegly


@dataclass
class DaneRZiS:
    """
    Rachunek Zysków i Strat (wariant porównawczy).
    Każda pozycja: (rok bieżący, rok ubiegły).

    Pozycje A-L zgodne ze standardem polskiej ustawy o rachunkowości:
      A. Przychody netto ze sprzedaży
      B. Koszty działalności operacyjnej
      C. Zysk (strata) ze sprzedaży = A - B
      D. Pozostałe przychody operacyjne
      E. Pozostałe koszty operacyjne
      F. Zysk (strata) z działalności operacyjnej = C + D - E
      G. Przychody finansowe
      H. Koszty finansowe
      I. Zysk (strata) brutto = F + G - H
      J. Podatek dochodowy
      K. Pozostałe obowiązkowe zmniejszenia zysku
      L. Zysk (strata) netto = I - J - K
    """
    przychody_sprzedazy:       Tuple[Decimal, Decimal] = (Decimal("0"), Decimal("0"))  # A
    koszty_operacyjne:         Tuple[Decimal, Decimal] = (Decimal("0"), Decimal("0"))  # B
    zysk_sprzedazy:            Tuple[Decimal, Decimal] = (Decimal("0"), Decimal("0"))  # C
    pozostale_przych_oper:     Tuple[Decimal, Decimal] = (Decimal("0"), Decimal("0"))  # D
    pozostale_koszty_oper:     Tuple[Decimal, Decimal] = (Decimal("0"), Decimal("0"))  # E
    zysk_dzialalnosci_oper:    Tuple[Decimal, Decimal] = (Decimal("0"), Decimal("0"))  # F
    przychody_finansowe:       Tuple[Decimal, Decimal] = (Decimal("0"), Decimal("0"))  # G
    koszty_finansowe:          Tuple[Decimal, Decimal] = (Decimal("0"), Decimal("0"))  # H
    zysk_brutto:               Tuple[Decimal, Decimal] = (Decimal("0"), Decimal("0"))  # I
    podatek_dochodowy:         Tuple[Decimal, Decimal] = (Decimal("0"), Decimal("0"))  # J
    pozostale_zmniejszenia:    Tuple[Decimal, Decimal] = (Decimal("0"), Decimal("0"))  # K
    zysk_netto:                Tuple[Decimal, Decimal] = (Decimal("0"), Decimal("0"))  # L

    # Mapa litera → pole (używane przy parsowaniu i walidacji)
    MAPA_POZYCJI = {
        "A": "przychody_sprzedazy",
        "B": "koszty_operacyjne",
        "C": "zysk_sprzedazy",
        "D": "pozostale_przych_oper",
        "E": "pozostale_koszty_oper",
        "F": "zysk_dzialalnosci_oper",
        "G": "przychody_finansowe",
        "H": "koszty_finansowe",
        "I": "zysk_brutto",
        "J": "podatek_dochodowy",
        "K": "pozostale_zmniejszenia",
        "L": "zysk_netto",
    }


@dataclass
class WyciagBankowy:
    """Pojedynczy wyciąg bankowy powiązany z analityką 130-X."""
    numer_konta_ksiegowego: str
    saldo_koncowe: Decimal
    rok_ostatniej_operacji: Optional[int] = None
    miesiac_ostatniej_operacji: Optional[int] = None
    numer_rachunku: str = ""
    bank_nazwa: str = ""
    wgrany_plik: bool = True

    @property
    def okres_opisowy(self) -> str:
        if self.miesiac_ostatniej_operacji and self.rok_ostatniej_operacji:
            return f"{MIESIACE_PL.get(self.miesiac_ostatniej_operacji, '?')} {self.rok_ostatniej_operacji}"
        elif self.rok_ostatniej_operacji:
            return str(self.rok_ostatniej_operacji)
        return "okres nieznany"


# =============================================================================
# FUNKCJE POMOCNICZE
# =============================================================================

def normalize_currency(wartosc: Union[str, float, int, None]) -> Decimal:
    """Normalizuje kwotę finansową do Decimal."""
    if wartosc is None:
        return Decimal("0")
    if isinstance(wartosc, (int, float)):
        try:
            return Decimal(str(round(float(wartosc), 2)))
        except (InvalidOperation, ValueError):
            return Decimal("0")
    tekst = str(wartosc).strip()
    if not tekst or tekst in ("-", "–", "0,00", "0.00"):
        return Decimal("0")
    tekst = re.sub(r"[złZŁPLN\xa0]", "", tekst).strip()
    if "," in tekst and "." in tekst:
        if tekst.rfind(",") > tekst.rfind("."):
            tekst = tekst.replace(".", "").replace(",", ".")
        else:
            tekst = tekst.replace(",", "")
    elif "," in tekst:
        czesci = tekst.split(",")
        if len(czesci) == 2 and len(czesci[1]) <= 2:
            tekst = tekst.replace(",", ".")
        else:
            tekst = tekst.replace(",", "")
    tekst = tekst.replace(" ", "")
    try:
        return Decimal(tekst).quantize(Decimal("0.01"))
    except InvalidOperation:
        raise ValueError(f"Nie można znormalizować kwoty: '{wartosc}'")


def normalize_konto(numer: str) -> str:
    if not numer:
        return ""
    return str(numer).strip().split("-")[0].strip()


def get_grupa(numer_konta: str) -> Optional[int]:
    syntetyka = normalize_konto(numer_konta)
    if syntetyka and syntetyka[0].isdigit():
        return int(syntetyka[0])
    return None


def wykryj_date_ostatniej_operacji(tekst: str) -> Optional[Tuple[int, int]]:
    """Wykrywa rok i miesiąc ostatniej operacji w wyciągu bankowym."""
    wszystkie_daty: List[date] = []
    wzorce = [
        r"(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})",
        r"(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})",
    ]
    for wzorzec in wzorce:
        for m in re.finditer(wzorzec, tekst):
            g = m.groups()
            try:
                if len(g[0]) == 4:
                    rok, mies, dzien = int(g[0]), int(g[1]), int(g[2])
                else:
                    dzien, mies, rok = int(g[0]), int(g[1]), int(g[2])
                if 1 <= mies <= 12 and 1 <= dzien <= 31 and 2000 <= rok <= 2100:
                    wszystkie_daty.append(date(rok, mies, dzien))
            except (ValueError, IndexError):
                continue
    for m in re.finditer(r"\b(\d{1,2})/(\d{4})\b", tekst):
        try:
            mies, rok = int(m.group(1)), int(m.group(2))
            if 1 <= mies <= 12 and 2000 <= rok <= 2100:
                wszystkie_daty.append(date(rok, mies, 1))
        except ValueError:
            continue
    if not wszystkie_daty:
        return None
    najpozniejsza = max(wszystkie_daty)
    return (najpozniejsza.year, najpozniejsza.month)


# =============================================================================
# PARSER ZOiS
# =============================================================================

class ParserZOiS:
    KOLUMNY_KONTO = ["konto", "numer konta", "nr konta", "symbol konta", "account"]
    KOLUMNY_NAZWA = ["nazwa", "nazwa konta", "opis", "name"]
    KOLUMNY_SALDO_WN = ["saldo wn", "saldo_wn", "debet", "wn", "saldo dt", "dt"]
    KOLUMNY_SALDO_MA = ["saldo ma", "saldo_ma", "kredyt", "ma", "saldo ct", "ct"]

    def _znajdz_kolumne(self, df, kandydaci):
        kol_lower = {k.lower().strip(): k for k in df.columns}
        for k in kandydaci:
            if k.lower() in kol_lower:
                return kol_lower[k.lower()]
        return None

    def parsuj_xlsx(self, dane_binarne: bytes) -> DaneZOiS:
        bufor = io.BytesIO(dane_binarne)
        try:
            wb = openpyxl.load_workbook(bufor, data_only=True)
        except Exception as e:
            raise ValueError(f"Błąd odczytu XLSX (ZOiS): {e}")
        arkusz = wb.active
        for nazwa in ["ZOiS", "Zestawienie", "ObrotyiSalda", "Sheet1"]:
            if nazwa in wb.sheetnames:
                arkusz = wb[nazwa]
                break
        wiersze = [list(w) for w in arkusz.iter_rows(values_only=True)]
        if not wiersze:
            raise ValueError("Arkusz ZOiS jest pusty.")
        idx = self._znajdz_wiersz_naglowka(wiersze)
        if idx is None:
            raise ValueError("Nie znaleziono wiersza nagłówkowego w ZOiS.")
        naglowki = [str(k).strip() if k else "" for k in wiersze[idx]]
        df = pd.DataFrame(wiersze[idx+1:], columns=naglowki)
        return self._parsuj_dataframe(df, "XLSX")

    def _znajdz_wiersz_naglowka(self, wiersze):
        kluczowe = {"konto", "saldo", "wn", "ma", "obroty", "numer"}
        for i, w in enumerate(wiersze[:20]):
            vals = {str(v).lower().strip() for v in w if v is not None}
            if len(vals & kluczowe) >= 2:
                return i
        return None

    # Regex do rozpoznawania wierszy w tekstowej wersji PDF Symfonii
    RE_NUMER_KONTA = re.compile(r"^(\d{2,3}(?:-\d+)*)\s+(.*)$")
    RE_LICZBA_PL = re.compile(r"-?\d+(?:\.\d{3})*,\d{1,2}")
    LINIE_IGNOROWANE = (
        "Suma strony", "Suma razem", "Z przeniesienia", "Do przeniesienia",
    )

    def parsuj_pdf(self, dane_binarne: bytes) -> DaneZOiS:
        """
        Parser PDF ZOiS z Symfonii.
        Symfonia eksportuje PDF jako tekst w kolumnach (bez obramowań tabeli),
        więc używamy ekstraktu tekstu + regex zamiast extract_tables().

        Strategia:
          1. Iteruj po liniach - każda potencjalnie to wiersz kontowy
          2. Wiersz zaczyna się od numeru konta (np. 130, 201-5, 403-1-1)
          3. Długie nazwy klientów mogą być wieloliniowe - obsługa przenoszenia
          4. Z końca linii bierzemy 8 liczb (BO Wn, BO Ma, obroty Wn/Ma,
             narast Wn/Ma, Saldo Wn, Saldo Ma) - interesują nas 2 ostatnie
        """
        if not PDF_AVAILABLE:
            raise ImportError("Brak biblioteki pdfplumber.")
        bufor = io.BytesIO(dane_binarne)
        linie = []
        with pdfplumber.open(bufor) as pdf:
            for strona in pdf.pages:
                t = dekoduj_cid(strona.extract_text() or "")
                linie.extend(t.splitlines())

        if not linie:
            raise ValueError("PDF jest pusty lub nie udało się ekstraktować tekstu.")

        wynik = DaneZOiS()
        czek_konto: Optional[str] = None
        czek_nazwa: str = ""

        def zapisz(numer: str, nazwa: str, liczby: List[str]):
            """
            Zapisuje pozycję do DaneZOiS - do syntetyk i/lub analityk.
            Z 8 liczb w wierszu bierze:
              - liczby[-2], liczby[-1]: saldo Wn, saldo Ma
              - liczby[-4], liczby[-3]: obroty narastająco Wn, Ma (do reguły 860)
            """
            try:
                wn = normalize_currency(liczby[-2])
                ma = normalize_currency(liczby[-1])
            except (ValueError, IndexError):
                return
            numer_syn = normalize_konto(numer)
            if not re.match(r"^\d+$", numer_syn):
                return

            # Wyciągnij obroty narastająco jeśli są dostępne
            narastajaco_wn = Decimal("0")
            narastajaco_ma = Decimal("0")
            if len(liczby) >= 8:
                try:
                    narastajaco_wn = normalize_currency(liczby[-4])
                    narastajaco_ma = normalize_currency(liczby[-3])
                except ValueError:
                    pass

            # Zapis jako analityka jeśli numer zawiera "-"
            if numer != numer_syn:
                wynik.konta_analityki[numer] = (wn, ma)
                if nazwa and numer not in wynik.opisy:
                    wynik.opisy[numer] = nazwa.strip()
            else:
                # Czysta syntetyka
                wynik.konta[numer_syn] = (wn, ma)
                wynik.obroty_narastajaco[numer_syn] = (narastajaco_wn, narastajaco_ma)
                if nazwa and numer_syn not in wynik.opisy:
                    wynik.opisy[numer_syn] = nazwa.strip()

        # Regex dla linii "Suma razem" – ostatnia linia ZOiS
        RE_SUMA_RAZEM = re.compile(r"^\s*Suma\s+razem\s+(.+)$", re.IGNORECASE)

        for linia in linie:
            ln = linia.strip()
            if not ln:
                continue

            # ── Wykrywanie linii "Suma razem" (fundamentalna kontrola Wn=Ma) ──
            m_suma = RE_SUMA_RAZEM.match(ln)
            if m_suma:
                liczby_suma = self.RE_LICZBA_PL.findall(m_suma.group(1))
                if len(liczby_suma) >= 8:
                    try:
                        wynik.sumy_razem = SumyRazemZOiS(
                            bo_wn=normalize_currency(liczby_suma[0]),
                            bo_ma=normalize_currency(liczby_suma[1]),
                            obroty_wn=normalize_currency(liczby_suma[2]),
                            obroty_ma=normalize_currency(liczby_suma[3]),
                            narastajaco_wn=normalize_currency(liczby_suma[4]),
                            narastajaco_ma=normalize_currency(liczby_suma[5]),
                            saldo_wn=normalize_currency(liczby_suma[6]),
                            saldo_ma=normalize_currency(liczby_suma[7]),
                            wykryto=True,
                        )
                    except ValueError:
                        pass
                continue

            # Pomijamy inne linie podsumowujące (Suma strony, Z przeniesienia, Do przeniesienia)
            if any(x in ln for x in self.LINIE_IGNOROWANE):
                continue

            m = self.RE_NUMER_KONTA.match(ln)
            if m:
                numer = m.group(1)
                reszta = m.group(2)
                liczby = self.RE_LICZBA_PL.findall(reszta)

                if len(liczby) >= 8:
                    # Cały wiersz w jednej linii
                    idx = reszta.find(liczby[0])
                    nazwa = reszta[:idx].strip() if idx > 0 else ""
                    zapisz(numer, nazwa, liczby)
                    czek_konto = None
                    czek_nazwa = ""
                else:
                    # Wiersz kontynuacyjny - nazwa/liczby w kolejnych liniach
                    czek_konto = numer
                    czek_nazwa = reszta
            elif czek_konto:
                liczby = self.RE_LICZBA_PL.findall(ln)
                if len(liczby) >= 8:
                    # Linia z liczbami dla oczekującego konta
                    zapisz(czek_konto, czek_nazwa, liczby)
                    czek_konto = None
                    czek_nazwa = ""
                else:
                    # Kontynuacja nazwy klienta
                    czek_nazwa += " " + ln

        # Dla kompatybilności: jeśli mamy syntetykę 130 ale nie analitykę, dodaj ją
        if "130" in wynik.konta and not any(
            normalize_konto(k) == "130" for k in wynik.konta_analityki
        ):
            wynik.konta_analityki["130"] = wynik.konta["130"]

        logger.info(
            f"ZOiS (PDF): {len(wynik.konta)} syntetyk, "
            f"{len(wynik.konta_analityki)} analityk."
        )
        return wynik

    def _parsuj_dataframe(self, df: pd.DataFrame, zrodlo: str) -> DaneZOiS:
        kol_konto = self._znajdz_kolumne(df, self.KOLUMNY_KONTO)
        kol_nazwa = self._znajdz_kolumne(df, self.KOLUMNY_NAZWA)
        kol_wn    = self._znajdz_kolumne(df, self.KOLUMNY_SALDO_WN)
        kol_ma    = self._znajdz_kolumne(df, self.KOLUMNY_SALDO_MA)
        brakujace = []
        if not kol_konto: brakujace.append("Konto")
        if not kol_wn:    brakujace.append("Saldo Wn")
        if not kol_ma:    brakujace.append("Saldo Ma")
        if brakujace:
            raise ValueError(
                f"Brak kolumn ZOiS ({zrodlo}): {', '.join(brakujace)}. "
                f"Dostępne: {list(df.columns)}"
            )
        wynik = DaneZOiS()
        for _, w in df.iterrows():
            numer_raw = w.get(kol_konto)
            if pd.isna(numer_raw) or not str(numer_raw).strip():
                continue
            numer_pelny = str(numer_raw).strip()
            numer_syn = normalize_konto(numer_pelny)
            if numer_syn.lower() in {"razem", "suma", "ogółem", "total"}:
                continue
            if not re.match(r"^\d+", numer_syn):
                continue
            try:
                wn = normalize_currency(w.get(kol_wn))
                ma = normalize_currency(w.get(kol_ma))
            except ValueError as e:
                logger.warning(f"Pominięto {numer_pelny}: {e}")
                continue
            if numer_pelny != numer_syn:
                iwn, ima = wynik.konta_analityki.get(numer_pelny, (Decimal("0"), Decimal("0")))
                wynik.konta_analityki[numer_pelny] = (iwn + wn, ima + ma)
                if kol_nazwa:
                    opis = w.get(kol_nazwa)
                    if not pd.isna(opis) and numer_pelny not in wynik.opisy:
                        wynik.opisy[numer_pelny] = str(opis).strip()
            iwn, ima = wynik.konta.get(numer_syn, (Decimal("0"), Decimal("0")))
            wynik.konta[numer_syn] = (iwn + wn, ima + ma)
            if kol_nazwa and numer_syn not in wynik.opisy:
                opis = w.get(kol_nazwa)
                if not pd.isna(opis):
                    wynik.opisy[numer_syn] = str(opis).strip()
        if "130" in wynik.konta and not any(
            normalize_konto(k) == "130" for k in wynik.konta_analityki
        ):
            wynik.konta_analityki["130"] = wynik.konta["130"]
        logger.info(
            f"ZOiS ({zrodlo}): {len(wynik.konta)} syntetyk, "
            f"{len(wynik.konta_analityki)} analityk."
        )
        return wynik


# =============================================================================
# PARSER BILANSU (v3.0 – pełna obsługa formatu Symfonii)
# =============================================================================

class ParserBilansu:
    """
    Parser Bilansu z Symfonii (wariant pełny).

    Wykrywa:
      - Sekcje AKTYWA i PASYWA (separator)
      - Pozycje A, B, C, D (aktywa) i A, B (pasywa)
      - Suma (pojawia się 2 razy: suma aktywów i suma pasywów)
      - Wynik netto (pasywa A.VI)
    """

    # Regex: pozycja A-D + tekst + 3 kwoty (bieżący, ubiegły, różnica)
    RE_POZ = re.compile(
        rf"^\s*(?:[+\-*]\s+)?([A-D])\s+.+?\s+({RE_KWOTA})\s+({RE_KWOTA})\s+({RE_KWOTA})\s*$"
    )
    # Regex: "Suma" + 3 kwoty
    RE_SUMA = re.compile(
        rf"^\s*Suma\s+({RE_KWOTA})\s+({RE_KWOTA})\s+({RE_KWOTA})\s*$"
    )
    # Regex: pozycja VI "Zysk (strata) netto" w kapitale własnym pasywów
    RE_WYNIK_NETTO = re.compile(
        rf"VI\s+Zysk.+?netto\s+({RE_KWOTA})\s+({RE_KWOTA})\s+({RE_KWOTA})"
    )

    def parsuj_xlsx(self, dane_binarne: bytes) -> DaneBilansu:
        """
        Parser XLSX dla Bilansu (wariant pełny).

        Symfonia eksportuje XLSX ze scalonymi komórkami gdzie litera pozycji
        (A-D) jest w jednej komórce, a liczby w dalekich kolumnach, z pustymi
        komórkami pomiędzy. Klasyczne " ".join() zamienia to na ciąg trudny
        do sparsowania regexem.

        Strategia: iterujemy po każdym wierszu i:
          1. Wykrywamy sekcję (AKTYWA / PASYWA)
          2. W wierszu szukamy komórki z pojedynczą literą A-D (pozycja główna)
             lub oznaczeniem "VI" (wynik netto A.VI pasywów) lub słowem "Suma"
          3. Bierzemy wszystkie liczbowe wartości z kolejnych komórek
          4. Z układu "bieżący | odchylenie | ubiegły | odchylenie | różnica"
             bierzemy liczby [0] i [2]
        """
        bufor = io.BytesIO(dane_binarne)
        try:
            wb = openpyxl.load_workbook(bufor, data_only=True)
        except Exception as e:
            raise ValueError(f"Błąd odczytu XLSX (Bilans): {e}")

        wynik = DaneBilansu()
        tryb: Optional[str] = None  # "A" = aktywa, "P" = pasywa
        znalezione_aktywa: Dict[str, Tuple[Decimal, Decimal]] = {}
        znalezione_pasywa: Dict[str, Tuple[Decimal, Decimal]] = {}
        sumy: List[Tuple[str, Decimal, Decimal]] = []

        def wyciagnij_liczby(kom_list) -> List[Decimal]:
            """Z listy komórek zwraca listę liczb (pomija puste i teksty)."""
            liczby: List[Decimal] = []
            for kom in kom_list:
                if kom is None or kom == "":
                    continue
                if isinstance(kom, (int, float)):
                    try:
                        liczby.append(Decimal(str(round(float(kom), 2))))
                    except (ValueError, InvalidOperation):
                        pass
                else:
                    tekst_kom = str(kom).strip()
                    if re.match(r"^-?[\d\s.,]+$", tekst_kom):
                        try:
                            liczby.append(normalize_currency(tekst_kom))
                        except ValueError:
                            pass
            return liczby

        def bierz_biezacy_ubiegly(liczby: List[Decimal]) -> Optional[Tuple[Decimal, Decimal]]:
            """Z listy liczb wyciąga (bieżący, ubiegły) pomijając odchylenia."""
            if len(liczby) >= 3:
                return (liczby[0], liczby[2])
            if len(liczby) >= 2:
                return (liczby[0], liczby[1])
            return None

        for w in wb.active.iter_rows(values_only=True):
            # ── Wykrywanie sekcji AKTYWA/PASYWA ──────────────────────────────
            # Nagłówek sekcji to wiersz w którym jest TYLKO słowo AKTYWA lub PASYWA
            # (reszta komórek pusta). W samej treści pozycji też występują słowa
            # typu "Aktywa trwałe", "Aktywa obrotowe" – nie mogą one przesuwać
            # trybu bo wtedy gubimy pozycje.
            niepuste = [str(k).strip() for k in w if k is not None and str(k).strip()]
            if niepuste == ["AKTYWA"]:
                tryb = "A"
                continue
            if niepuste == ["PASYWA"]:
                tryb = "P"
                continue

            # ── "Suma" – pojawia się 2x (aktywa, pasywa) ─────────────────────
            for i, kom in enumerate(w):
                if kom is None:
                    continue
                if str(kom).strip().lower() == "suma":
                    liczby = wyciagnij_liczby(w[i + 1:])
                    para = bierz_biezacy_ubiegly(liczby)
                    if para and tryb:
                        sumy.append((tryb, para[0], para[1]))
                    break

            # ── Wynik netto (A.VI pasywów) – kolumna z "VI" + tekst "Zysk" ──
            # Szukamy wiersza który zawiera zarówno "VI" jako osobną komórkę
            # jak i słowo "Zysk" w innej komórce
            if tryb == "P":
                ma_VI = False
                ma_zysk = False
                poz_VI = -1
                for i, kom in enumerate(w):
                    if kom is None:
                        continue
                    tekst = str(kom).strip()
                    if tekst == "VI":
                        ma_VI = True
                        poz_VI = i
                    elif "zysk" in tekst.lower() and "netto" in tekst.lower():
                        ma_zysk = True
                if ma_VI and ma_zysk and poz_VI >= 0:
                    liczby = wyciagnij_liczby(w[poz_VI + 1:])
                    para = bierz_biezacy_ubiegly(liczby)
                    if para:
                        wynik.wynik_netto_biezacy = para[0]
                        wynik.wynik_netto_ubiegly = para[1]

            # ── Pozycja główna A/B/C/D ────────────────────────────────────────
            if tryb:
                for i, kom in enumerate(w):
                    if kom is None:
                        continue
                    tekst = str(kom).strip()
                    if len(tekst) == 1 and tekst in "ABCD":
                        liczby = wyciagnij_liczby(w[i + 1:])
                        para = bierz_biezacy_ubiegly(liczby)
                        if para:
                            target = znalezione_aktywa if tryb == "A" else znalezione_pasywa
                            # "first wins" – pierwsza pozycja główna A/B/C/D
                            # (kolejne np. "A.I" to podpozycje których nie
                            # chcemy nadpisywać)
                            if tekst not in target:
                                target[tekst] = para
                        break  # jedna pozycja na wiersz

        # ── Przepisanie do DaneBilansu ───────────────────────────────────────
        if "A" in znalezione_aktywa:
            wynik.aktywa_trwale_biezacy, wynik.aktywa_trwale_ubiegly = znalezione_aktywa["A"]
        if "B" in znalezione_aktywa:
            wynik.aktywa_obrotowe_biezacy, wynik.aktywa_obrotowe_ubiegly = znalezione_aktywa["B"]
        if "C" in znalezione_aktywa:
            wynik.nalezne_wplaty_biezacy, wynik.nalezne_wplaty_ubiegly = znalezione_aktywa["C"]
        if "D" in znalezione_aktywa:
            wynik.udzialy_wlasne_biezacy, wynik.udzialy_wlasne_ubiegly = znalezione_aktywa["D"]

        if "A" in znalezione_pasywa:
            wynik.kapital_wlasny_biezacy, wynik.kapital_wlasny_ubiegly = znalezione_pasywa["A"]
        if "B" in znalezione_pasywa:
            wynik.zobowiazania_biezacy, wynik.zobowiazania_ubiegly = znalezione_pasywa["B"]

        # Sumy – pierwsza z trybu A (aktywa), druga z P (pasywa)
        for tr, b, u in sumy:
            if tr == "A" and wynik.suma_aktywow_biezacy == Decimal("0"):
                wynik.suma_aktywow_biezacy = b
                wynik.suma_aktywow_ubiegly = u
            elif tr == "P" and wynik.suma_pasywow_biezacy == Decimal("0"):
                wynik.suma_pasywow_biezacy = b
                wynik.suma_pasywow_ubiegly = u

        logger.info(
            f"Bilans (XLSX): Aktywa={wynik.suma_aktywow_biezacy:,.2f}, "
            f"Pasywa={wynik.suma_pasywow_biezacy:,.2f}, "
            f"Wynik netto={wynik.wynik_netto_biezacy:,.2f}"
        )
        return wynik

    def parsuj_pdf(self, dane_binarne: bytes) -> DaneBilansu:
        """PDF – ekstrahuje tekst i parsuje linia po linii."""
        if not PDF_AVAILABLE:
            raise ImportError("Brak pdfplumber.")
        bufor = io.BytesIO(dane_binarne)
        wszystkie = []
        with pdfplumber.open(bufor) as pdf:
            for strona in pdf.pages:
                t = dekoduj_cid(strona.extract_text() or "")
                wszystkie.extend(t.splitlines())
        return self._parsuj_linie(wszystkie)

    def _parsuj_linie(self, linie: List[str]) -> DaneBilansu:
        """
        Główna logika parsowania. Iteruje po liniach śledząc w której sekcji
        (AKTYWA/PASYWA) się znajdujemy.
        """
        wynik = DaneBilansu()
        tryb: Optional[str] = None  # "A" = aktywa, "P" = pasywa
        sumy: List[Tuple[str, Decimal, Decimal]] = []  # [(tryb, bieżący, ubiegły)]
        znalezione_aktywa: Dict[str, Tuple[Decimal, Decimal]] = {}
        znalezione_pasywa: Dict[str, Tuple[Decimal, Decimal]] = {}

        for linia in linie:
            # ── Separator sekcji ─────────────────────────────────────────────
            if "AKTYWA" in linia and "PASYWA" not in linia:
                tryb = "A"
                continue
            if "PASYWA" in linia:
                tryb = "P"
                continue

            # ── Wynik netto (A.VI pasywów) ──────────────────────────────────
            if tryb == "P":
                m_wn = self.RE_WYNIK_NETTO.search(linia)
                if m_wn:
                    try:
                        wynik.wynik_netto_biezacy = normalize_currency(m_wn.group(1))
                        wynik.wynik_netto_ubiegly = normalize_currency(m_wn.group(2))
                    except ValueError:
                        pass

            # ── Suma (pojawia się 2x: aktywa i pasywa) ───────────────────────
            m_suma = self.RE_SUMA.match(linia)
            if m_suma and tryb:
                try:
                    b = normalize_currency(m_suma.group(1))
                    u = normalize_currency(m_suma.group(2))
                    sumy.append((tryb, b, u))
                except ValueError:
                    pass
                continue

            # ── Pozycja A/B/C/D w aktywach lub A/B w pasywach ────────────────
            m = self.RE_POZ.match(linia)
            if m and tryb:
                lit = m.group(1)
                try:
                    b = normalize_currency(m.group(2))
                    u = normalize_currency(m.group(3))
                except ValueError:
                    continue

                target = znalezione_aktywa if tryb == "A" else znalezione_pasywa
                # "First wins" – pierwsza pozycja wiąże się z główną, kolejne ignorujemy
                if lit not in target:
                    target[lit] = (b, u)

        # ── Przepisanie do struktury DaneBilansu ─────────────────────────────
        if "A" in znalezione_aktywa:
            wynik.aktywa_trwale_biezacy, wynik.aktywa_trwale_ubiegly = znalezione_aktywa["A"]
        if "B" in znalezione_aktywa:
            wynik.aktywa_obrotowe_biezacy, wynik.aktywa_obrotowe_ubiegly = znalezione_aktywa["B"]
        if "C" in znalezione_aktywa:
            wynik.nalezne_wplaty_biezacy, wynik.nalezne_wplaty_ubiegly = znalezione_aktywa["C"]
        if "D" in znalezione_aktywa:
            wynik.udzialy_wlasne_biezacy, wynik.udzialy_wlasne_ubiegly = znalezione_aktywa["D"]

        if "A" in znalezione_pasywa:
            wynik.kapital_wlasny_biezacy, wynik.kapital_wlasny_ubiegly = znalezione_pasywa["A"]
        if "B" in znalezione_pasywa:
            wynik.zobowiazania_biezacy, wynik.zobowiazania_ubiegly = znalezione_pasywa["B"]

        # Sumy: pierwsza to aktywa, druga to pasywa
        for tr, b, u in sumy:
            if tr == "A" and wynik.suma_aktywow_biezacy == Decimal("0"):
                wynik.suma_aktywow_biezacy = b
                wynik.suma_aktywow_ubiegly = u
            elif tr == "P" and wynik.suma_pasywow_biezacy == Decimal("0"):
                wynik.suma_pasywow_biezacy = b
                wynik.suma_pasywow_ubiegly = u

        logger.info(
            f"Bilans: Aktywa={wynik.suma_aktywow_biezacy:,.2f}, "
            f"Pasywa={wynik.suma_pasywow_biezacy:,.2f}, "
            f"Wynik netto={wynik.wynik_netto_biezacy:,.2f}"
        )
        return wynik


# =============================================================================
# PARSER RACHUNKU ZYSKÓW I STRAT (v3.0 – nowość)
# =============================================================================

class ParserRZiS:
    """
    Parser Rachunku Zysków i Strat (wariant porównawczy).

    Strategia: "last wins" – iteruje po liniach i nadpisuje słownik pozycji.
    Dzięki temu:
      - Litery A-H, J, K, L są unikalne (pierwsza = główna)
      - Litera "I" występuje wielokrotnie jako rzymska cyfra I w podsekcjach
        (np. "I Amortyzacja" w sekcji B), ale OSTATNIA "I" = główna (Zysk brutto)
    """

    RE_POZ = re.compile(
        rf"^\s*(?:[+\-*]\s+)?([A-L])\s+.+?\s+({RE_KWOTA})\s+({RE_KWOTA})\s*$"
    )

    def parsuj_xlsx(self, dane_binarne: bytes) -> DaneRZiS:
        """
        Parser XLSX dla RZiS (wariant porównawczy).

        Symfonia eksportuje XLSX ze scalonymi komórkami gdzie litera pozycji
        (A-L) jest w jednej komórce, a liczby w dalekich kolumnach, z pustymi
        komórkami pomiędzy. Klasyczne " ".join() zamienia to na ciąg trudny
        do sparsowania regexem.

        Strategia: iterujemy po każdym wierszu i:
          1. Znajdujemy komórkę zawierającą dokładnie pojedynczą literę A-L
          2. Bierzemy wszystkie liczbowe wartości z kolejnych komórek
          3. Pierwsze dwie liczby to: rok bieżący i rok ubiegły
             (trzecia to odchylenie, czwarta różnica – ignorujemy)

        Strategia "last wins" dla pozycji I (rzymska I w podsekcjach B, G, H
        kolidowałaby z główną literą I – ostatnie wystąpienie wygrywa).
        """
        bufor = io.BytesIO(dane_binarne)
        try:
            wb = openpyxl.load_workbook(bufor, data_only=True)
        except Exception as e:
            raise ValueError(f"Błąd odczytu XLSX (RZiS): {e}")

        znalezione: Dict[str, Tuple[Decimal, Decimal]] = {}

        for w in wb.active.iter_rows(values_only=True):
            # Znajdź komórkę będącą dokładnie pojedynczą literą A-L
            poz_litery = None
            for i, kom in enumerate(w):
                if kom is None:
                    continue
                tekst = str(kom).strip()
                if len(tekst) == 1 and tekst in "ABCDEFGHIJKL":
                    poz_litery = i
                    break

            if poz_litery is None:
                # Alternatywnie – zwykły format PDF ("A  Tekst  kwota kwota")
                # może być używany jeśli plik był konwertowany. Spróbujmy regexu.
                linia = " ".join(str(k) for k in w if k is not None)
                m = self.RE_POZ.match(linia)
                if m:
                    lit = m.group(1)
                    try:
                        b = normalize_currency(m.group(2))
                        u = normalize_currency(m.group(3))
                        znalezione[lit] = (b, u)
                    except ValueError:
                        pass
                continue

            litera = str(w[poz_litery]).strip()

            # Od tej pozycji zbieraj wszystkie liczby
            liczby: List[Decimal] = []
            for kom in w[poz_litery + 1:]:
                if kom is None or kom == "":
                    continue
                # Liczba jako float/int lub jako string z polskim formatem
                if isinstance(kom, (int, float)):
                    try:
                        liczby.append(Decimal(str(round(float(kom), 2))))
                    except (ValueError, InvalidOperation):
                        pass
                else:
                    tekst_kom = str(kom).strip()
                    # Pomijamy teksty które nie wyglądają jak liczba
                    if re.match(r"^-?[\d\s.,]+$", tekst_kom):
                        try:
                            liczby.append(normalize_currency(tekst_kom))
                        except ValueError:
                            pass

            if len(liczby) >= 2:
                # Układ kolumn w XLSX Symfonii:
                #   [0] rok bieżący
                #   [1] odchylenie (zawsze 0)
                #   [2] rok ubiegły
                #   [3] odchylenie (zawsze 0)
                #   [4] różnica
                # Jeśli mamy >= 3 liczby – bierzemy [0] i [2] (omijając odchylenie)
                # Fallback – gdy kolumn jest tylko 2 (format skrócony) bierzemy [0] i [1]
                if len(liczby) >= 3:
                    znalezione[litera] = (liczby[0], liczby[2])
                else:
                    znalezione[litera] = (liczby[0], liczby[1])

        wynik = DaneRZiS()
        for lit, pole in DaneRZiS.MAPA_POZYCJI.items():
            if lit in znalezione:
                setattr(wynik, pole, znalezione[lit])

        logger.info(
            f"RZiS (XLSX): Przychody(A)={wynik.przychody_sprzedazy[0]:,.2f}, "
            f"Koszty(B)={wynik.koszty_operacyjne[0]:,.2f}, "
            f"Zysk netto(L)={wynik.zysk_netto[0]:,.2f}"
        )
        return wynik

    def parsuj_pdf(self, dane_binarne: bytes) -> DaneRZiS:
        if not PDF_AVAILABLE:
            raise ImportError("Brak pdfplumber.")
        bufor = io.BytesIO(dane_binarne)
        linie = []
        with pdfplumber.open(bufor) as pdf:
            for strona in pdf.pages:
                t = dekoduj_cid(strona.extract_text() or "")
                linie.extend(t.splitlines())
        return self._parsuj_linie(linie)

    def _parsuj_linie(self, linie: List[str]) -> DaneRZiS:
        znalezione: Dict[str, Tuple[Decimal, Decimal]] = {}

        for linia in linie:
            m = self.RE_POZ.match(linia)
            if m:
                lit = m.group(1)
                try:
                    b = normalize_currency(m.group(2))
                    u = normalize_currency(m.group(3))
                    # "last wins" – ostatnie wystąpienie nadpisuje
                    znalezione[lit] = (b, u)
                except ValueError:
                    continue

        wynik = DaneRZiS()
        for lit, pole in DaneRZiS.MAPA_POZYCJI.items():
            if lit in znalezione:
                setattr(wynik, pole, znalezione[lit])

        logger.info(
            f"RZiS: Przychody(A)={wynik.przychody_sprzedazy[0]:,.2f}, "
            f"Koszty(B)={wynik.koszty_operacyjne[0]:,.2f}, "
            f"Zysk netto(L)={wynik.zysk_netto[0]:,.2f}"
        )
        return wynik


# =============================================================================
# PARSER WYCIĄGU BANKOWEGO / ZESTAWIENIA OPERACJI (v3.3)
# =============================================================================
# Obsługuje oba typy dokumentów:
#   - Standardowy wyciąg bankowy (np. Santander)
#   - Elektroniczne zestawienie operacji (np. mBank)
# Radzi sobie z tekstem bez spacji (pdfplumber przy niektórych fontach zwraca
# "Saldokońcowe:20992,02" zamiast "Saldo końcowe: 20 992,02").
# =============================================================================

class ParserWyciaguBankowego:
    # Wzorce dla salda końcowego. Niektóre banki (np. mBank) eksportują PDF
    # w którym pdfplumber zwraca tekst bez spacji ("Saldokońcowe:20992,02"),
    # dlatego szukamy zarówno z jak i bez spacji.
    KLUCZE_SALDO = [
        "saldo końcowe", "saldokońcowe",
        "saldo na koniec", "saldonakoniec",
        "closing balance", "closingbalance",
        "stan na koniec", "stannakoniec",
        "saldo końca okresu", "saldokońcaokresu",
        "saldo po operacji",  # ostatni wiersz tabeli transakcji
    ]
    WZORZEC_IBAN = re.compile(r"(?:PL)?\s*(\d{2}\s*(?:\d{4}\s*){6})")
    # Regex dla kwoty – również tolerancyjny na brak spacji
    RE_SALDO_KWOTA = re.compile(r"(-?\d{1,3}(?:[\s.]\d{3})*,\d{2}|-?\d+,\d{2})")
    # Regex dla okresu zestawienia: "za okres od 2025-12-01 do 2025-12-31"
    RE_OKRES = re.compile(
        r"(?:okres|zaokres)[^\d]*?(\d{4})-(\d{1,2})-(\d{1,2})[^\d]+?(\d{4})-(\d{1,2})-(\d{1,2})",
        re.IGNORECASE,
    )
    ZNANE_BANKI = [
        "Santander", "PKO BP", "PKO Bank", "mBank", "ING Bank", "Pekao",
        "BNP Paribas", "Millennium", "Credit Agricole", "Alior", "Citi",
        "Getin", "Nest Bank", "Raiffeisen", "Deutsche Bank", "BOŚ",
    ]

    def parsuj(self, dane_binarne: bytes, format_pliku: str):
        if format_pliku.lower() == "pdf":
            return self._parsuj_pdf(dane_binarne)
        return self._parsuj_xlsx(dane_binarne)

    def _parsuj_pdf(self, dane_binarne: bytes):
        if not PDF_AVAILABLE:
            raise ImportError("Brak pdfplumber.")
        bufor = io.BytesIO(dane_binarne)
        saldo = Decimal("0")
        caly_tekst = ""
        with pdfplumber.open(bufor) as pdf:
            for strona in pdf.pages:
                caly_tekst += dekoduj_cid(strona.extract_text() or "") + "\n"
            # Szukamy salda od końca (ostatnia strona = saldo końcowe)
            for strona in reversed(pdf.pages):
                t = dekoduj_cid(strona.extract_text() or "")
                s = self._wyciagnij_saldo(t)
                if s and s != Decimal("0"):
                    saldo = s
                    break

        # Preferuj datę z nagłówka "za okres od X do Y" – pewniejsza niż
        # szukanie najpóźniejszej daty w całym tekście (mBank ma np. datę
        # "następnej kapitalizacji 2026-01-31" w nagłówku, która nie powinna
        # być traktowana jako data ostatniej operacji).
        rok, mies = self._wykryj_okres_zestawienia(caly_tekst)
        if rok is None:
            data_op = wykryj_date_ostatniej_operacji(caly_tekst)
            rok, mies = data_op if data_op else (None, None)

        return saldo, rok, mies, self._iban(caly_tekst), self._bank(caly_tekst)

    def _parsuj_xlsx(self, dane_binarne: bytes):
        bufor = io.BytesIO(dane_binarne)
        try:
            df = pd.read_excel(bufor, engine="openpyxl", header=None)
        except Exception as e:
            raise ValueError(f"Błąd odczytu XLSX wyciągu: {e}")
        saldo = Decimal("0")
        caly_tekst = ""
        for _, wiersz in df.iterrows():
            for kom in wiersz:
                if isinstance(kom, str):
                    caly_tekst += kom + " "
                    if any(k in kom.lower() for k in self.KLUCZE_SALDO):
                        for v in wiersz:
                            try:
                                kw = normalize_currency(v)
                                if kw != Decimal("0"):
                                    saldo = kw
                                    break
                            except (ValueError, TypeError):
                                pass
                elif kom is not None:
                    caly_tekst += str(kom) + " "
        data_op = wykryj_date_ostatniej_operacji(caly_tekst)
        rok, mies = data_op if data_op else (None, None)
        return saldo, rok, mies, self._iban(caly_tekst), self._bank(caly_tekst)

    def _wyciagnij_saldo(self, tekst: str) -> Optional[Decimal]:
        """
        Szuka salda końcowego w tekście wyciągu.

        Radzi sobie z dwoma formatami:
          - Ze spacjami: "Saldo końcowe: 20.992,02"
          - Bez spacji (mBank, pdfplumber bez spacji): "Saldokońcowe:20992,02"

        Priorytet: ostatnie wystąpienie słowa kluczowego w tekście (tak jak
        na wyciągu – saldo końcowe jest na dole ostatniej strony).
        """
        # Znormalizuj: lowercase + usuń spacje dla porównania
        tekst_no_space = re.sub(r"\s+", "", tekst).lower()
        tekst_lower = tekst.lower()

        # Spróbuj najpierw metodą "linia po linii" (klasyczny format)
        for linia in reversed(tekst.splitlines()):
            ll = linia.lower()
            if any(k in ll for k in self.KLUCZE_SALDO if " " in k):
                # Standardowy format ze spacjami – szukaj kwot w linii
                kwoty = self.RE_SALDO_KWOTA.findall(linia)
                for kwota_str in reversed(kwoty):
                    try:
                        kw = normalize_currency(kwota_str)
                        if kw != Decimal("0"):
                            return kw
                    except (ValueError, TypeError):
                        pass

        # Fallback: tekst bez spacji (mBank)
        for klucz in self.KLUCZE_SALDO:
            klucz_clean = klucz.replace(" ", "")
            # Szukaj OSTATNIEGO wystąpienia (saldo końcowe = na końcu dokumentu)
            idx = tekst_no_space.rfind(klucz_clean)
            if idx == -1:
                continue
            # Wyciągnij fragment po słowie kluczowym (do 50 znaków)
            po_kluczu = tekst_no_space[idx + len(klucz_clean):idx + len(klucz_clean) + 50]
            kwoty = self.RE_SALDO_KWOTA.findall(po_kluczu)
            for kwota_str in kwoty:
                try:
                    kw = normalize_currency(kwota_str)
                    if kw != Decimal("0"):
                        return kw
                except (ValueError, TypeError):
                    pass
        return None

    def _wykryj_okres_zestawienia(self, tekst: str) -> Tuple[Optional[int], Optional[int]]:
        """
        Wykrywa okres zestawienia z nagłówka dokumentu.
        Szuka frazy "za okres od YYYY-MM-DD do YYYY-MM-DD" i zwraca rok/miesiąc
        daty końcowej. To pewniejsze niż szukanie najpóźniejszej daty w tekście,
        bo niektóre banki zawierają daty "przyszłe" (np. data kapitalizacji).
        """
        # Usuń spacje żeby radzić sobie z różnymi formatami
        tekst_no_space = re.sub(r"\s+", "", tekst)
        m = self.RE_OKRES.search(tekst_no_space)
        if m:
            try:
                rok_do, mies_do = int(m.group(4)), int(m.group(5))
                if 2000 <= rok_do <= 2100 and 1 <= mies_do <= 12:
                    return rok_do, mies_do
            except (ValueError, IndexError):
                pass
        return None, None

    def _iban(self, tekst: str) -> str:
        m = self.WZORZEC_IBAN.search(tekst)
        return re.sub(r"\s+", "", m.group(1)) if m else ""

    def _bank(self, tekst: str) -> str:
        tl = tekst.lower()
        for bank in self.ZNANE_BANKI:
            if bank.lower() in tl:
                return bank
        return ""


# =============================================================================
# GŁÓWNA KLASA AUDYTORA
# =============================================================================

class SymfoniaYearEndAuditor:
    """
    Audytor jakości danych – zamknięcie roku v3.0.

    API:
        parsuj_zois(bytes, format)    → DaneZOiS
        parsuj_bilans(bytes, format)  → DaneBilansu
        parsuj_rzis(bytes, format)    → DaneRZiS
        parsuj_wyciag(konto, bytes, format) → WyciagBankowy
        check_accounting_logic(...)   → weryfikacja
        generate_audit_report(...)    → raport
        run_full_audit(...)           → wszystko w jednym wywołaniu
    """

    def __init__(self):
        self._parser_zois   = ParserZOiS()
        self._parser_bilans = ParserBilansu()
        self._parser_rzis   = ParserRZiS()
        self._parser_wyciag = ParserWyciaguBankowego()
        self._wyniki: List[PunktKontroli] = []

    # ── Publiczne API – parsowanie ──────────────────────────────────────────

    def parsuj_zois(self, dane_binarne: bytes, format_pliku: str = "xlsx") -> DaneZOiS:
        if format_pliku.lower() == "pdf":
            return self._parser_zois.parsuj_pdf(dane_binarne)
        return self._parser_zois.parsuj_xlsx(dane_binarne)

    def parsuj_bilans(self, dane_binarne: bytes, format_pliku: str = "xlsx") -> DaneBilansu:
        if format_pliku.lower() == "pdf":
            return self._parser_bilans.parsuj_pdf(dane_binarne)
        return self._parser_bilans.parsuj_xlsx(dane_binarne)

    def parsuj_rzis(self, dane_binarne: bytes, format_pliku: str = "xlsx") -> DaneRZiS:
        if format_pliku.lower() == "pdf":
            return self._parser_rzis.parsuj_pdf(dane_binarne)
        return self._parser_rzis.parsuj_xlsx(dane_binarne)

    def parsuj_wyciag(
        self,
        numer_konta_ksiegowego: str,
        dane_binarne: bytes,
        format_pliku: str = "pdf",
    ) -> WyciagBankowy:
        saldo, rok, mies, iban, bank = self._parser_wyciag.parsuj(
            dane_binarne, format_pliku
        )
        return WyciagBankowy(
            numer_konta_ksiegowego=numer_konta_ksiegowego,
            saldo_koncowe=saldo,
            rok_ostatniej_operacji=rok,
            miesiac_ostatniej_operacji=mies,
            numer_rachunku=iban,
            bank_nazwa=bank,
            wgrany_plik=True,
        )

    # ── Publiczne API – weryfikacja ─────────────────────────────────────────

    def check_accounting_logic(
        self,
        dane_zois:   Optional[DaneZOiS],
        dane_bilans: Optional[DaneBilansu]  = None,
        dane_rzis:   Optional[DaneRZiS]     = None,
        dane_krs:    Optional[DaneKRS]      = None,
        wyciagi:     Optional[List[WyciagBankowy]] = None,
        rok_obrachunkowy: int = 2024,
    ) -> List[PunktKontroli]:
        self._wyniki = []
        wyciagi = wyciagi or []

        # ── ZOiS – konta ────────────────────────────────────────────────────
        if dane_zois is None:
            self._wyniki.append(PunktKontroli(
                konto="ZOiS", punkt="Wczytanie ZOiS", status=StatusAudytu.BRAK,
                uwagi="Nie dostarczono pliku ZOiS.",
            ))
        else:
            # R2: Fundamentalna kontrola dwustronnego zapisu (Wn=Ma)
            self._weryfikuj_obroty_wn_ma(dane_zois)
            self._weryfikuj_konta_bankowe(dane_zois, wyciagi, rok_obrachunkowy)
            self._weryfikuj_konto_145(dane_zois)
            self._weryfikuj_konto_200(dane_zois)
            self._weryfikuj_konto_202(dane_zois)
            self._weryfikuj_konto_230(dane_zois)
            self._weryfikuj_konto_229(dane_zois)
            self._weryfikuj_konto_220(dane_zois)
            self._weryfikuj_konto_700(dane_zois)
            self._weryfikuj_grupe_4(dane_zois)
            # R1: Weryfikacja konta 860 (wymaga RZiS do porównania)
            self._weryfikuj_konto_860(dane_zois, dane_rzis)

        # ── Bilans ──────────────────────────────────────────────────────────
        self._weryfikuj_bilans(dane_bilans)

        # ── RZiS ────────────────────────────────────────────────────────────
        self._weryfikuj_rzis(dane_rzis)

        # ── Reguły krzyżowe ZOiS ↔ RZiS ↔ Bilans ────────────────────────────
        self._weryfikuj_krzyzowe(dane_zois, dane_bilans, dane_rzis)

        # ── R3: Weryfikacja KRS ────────────────────────────────────────────
        if dane_krs is not None:
            self._weryfikuj_krs(dane_zois, dane_bilans, dane_krs)

        return self._wyniki

    def generate_audit_report(
        self,
        wyniki: List[PunktKontroli],
        nazwa_podmiotu: str = "Podmiot",
        rok: int = 2024,
    ) -> Dict:
        linia = "═" * 70
        L = [
            linia,
            f"  RAPORT KONTROLI JAKOŚCI DANYCH – ZAMKNIĘCIE ROKU {rok}",
            f"  Podmiot: {nazwa_podmiotu}",
            f"  Wygenerowano przez: SymfoniaYearEndAuditor v3.3",
            linia, "",
            f"{'ŹRÓDŁO':<14} {'STATUS':<20} {'PUNKT KONTROLI':<42} UWAGI",
            "─" * 110,
        ]
        stats = {s: 0 for s in StatusAudytu}
        wyniki_slow = []
        for pkt in wyniki:
            stats[pkt.status] += 1
            wartosc = f" [{pkt.wartosc}]" if pkt.wartosc else ""
            L.append(
                f"{pkt.konto:<14} {pkt.status.value:<20} "
                f"{pkt.punkt:<42} {pkt.uwagi}{wartosc}"
            )
            wyniki_slow.append({
                "konto": pkt.konto, "status": pkt.status.value,
                "punkt": pkt.punkt, "uwagi": pkt.uwagi, "wartosc": pkt.wartosc,
            })

        bledy = [p for p in wyniki if p.status == StatusAudytu.BLAD]
        ostrz = [p for p in wyniki if p.status == StatusAudytu.OSTRZEZ]
        brak  = [p for p in wyniki if p.status == StatusAudytu.BRAK]

        L.extend(["", linia, "  PODSUMOWANIE", linia])
        L.append(f"  ✅ OK:            {stats[StatusAudytu.OK]}")
        L.append(f"  ❌ BŁĘDY:         {stats[StatusAudytu.BLAD]}")
        L.append(f"  ⚠️  OSTRZEŻENIA:  {stats[StatusAudytu.OSTRZEZ]}")
        L.append(f"  🔍 BRAK DANYCH:  {stats[StatusAudytu.BRAK]}")

        if bledy:
            L.extend(["", "  ❌ WYMAGANE DZIAŁANIA (BŁĘDY KRYTYCZNE):"])
            for p in bledy:
                L.append(f"    → {p.konto}: {p.uwagi}")
        if ostrz:
            L.extend(["", "  ⚠️  DO WYJAŚNIENIA (OSTRZEŻENIA):"])
            for p in ostrz:
                L.append(f"    → {p.konto}: {p.uwagi}")
        if brak:
            L.extend(["", "  🔍 BRAKI DANYCH:"])
            for p in brak:
                L.append(f"    → {p.konto}: {p.uwagi}")

        L.append("")
        if stats[StatusAudytu.BLAD] == 0 and stats[StatusAudytu.OSTRZEZ] == 0:
            L.append("  🎉 OCENA KOŃCOWA: DANE SPÓJNE – gotowe do badania.")
        elif stats[StatusAudytu.BLAD] == 0:
            L.append("  🟡 OCENA KOŃCOWA: WYMAGA WYJAŚNIENIA.")
        else:
            L.append("  🔴 OCENA KOŃCOWA: DANE NIESPÓJNE – wymagane korekty.")
        L.append(linia)

        return {
            "tekst": "\n".join(L),
            "wyniki": wyniki_slow,
            "podsumowanie": {
                "ok": stats[StatusAudytu.OK],
                "bledy": stats[StatusAudytu.BLAD],
                "ostrzezenia": stats[StatusAudytu.OSTRZEZ],
                "brak_danych": stats[StatusAudytu.BRAK],
                "podmiot": nazwa_podmiotu,
                "rok": rok,
            },
        }

    def run_full_audit(
        self,
        *,
        zois_bytes:   Optional[bytes] = None,
        zois_format:  str = "xlsx",
        bilans_bytes: Optional[bytes] = None,
        bilans_format: str = "xlsx",
        rzis_bytes:   Optional[bytes] = None,
        rzis_format:  str = "pdf",
        wyciagi: Optional[List[WyciagBankowy]] = None,
        nazwa_podmiotu: str = "Podmiot",
        rok: int = 2024,
    ) -> Dict:
        dane_zois   = self.parsuj_zois(zois_bytes, zois_format) if zois_bytes else None
        dane_bilans = self.parsuj_bilans(bilans_bytes, bilans_format) if bilans_bytes else None
        dane_rzis   = self.parsuj_rzis(rzis_bytes, rzis_format) if rzis_bytes else None

        wyniki = self.check_accounting_logic(
            dane_zois, dane_bilans, dane_rzis, wyciagi or [],
            rok_obrachunkowy=rok,
        )
        return self.generate_audit_report(wyniki, nazwa_podmiotu, rok)

    # ─── WERYFIKACJA BANK ─────────────────────────────────────────────────────

    def _weryfikuj_konta_bankowe(
        self,
        dane_zois: DaneZOiS,
        wyciagi: List[WyciagBankowy],
        rok_obrachunkowy: int,
    ) -> None:
        konta_bankowe = dane_zois.pobierz_konta_bankowe()
        if not konta_bankowe:
            self._wyniki.append(PunktKontroli(
                konto="130", punkt="Rachunki bankowe",
                status=StatusAudytu.BRAK,
                uwagi="CRITICAL: Nie odnaleziono konta 130 w ZOiS.",
            ))
            return

        mapa = {w.numer_konta_ksiegowego: w for w in wyciagi}
        self._wyniki.append(PunktKontroli(
            konto="130", punkt="Wykryte rachunki bankowe",
            status=StatusAudytu.INFO,
            uwagi=f"Wykryto {len(konta_bankowe)} rachunek(ów) w ZOiS, wgrano {len(wyciagi)} wyciąg(ów).",
        ))

        rachunki_bez_wyciagu = []
        for numer_ks, opis, saldo_zois in konta_bankowe:
            wyciag = mapa.get(numer_ks)
            if wyciag is None:
                rachunki_bez_wyciagu.append((numer_ks, opis, saldo_zois))
                self._wyniki.append(PunktKontroli(
                    konto=numer_ks, punkt=f"Rachunek {opis}",
                    status=StatusAudytu.BRAK,
                    uwagi=f"Brak wgranego wyciągu. Saldo ZOiS: {saldo_zois:,.2f} zł.",
                ))
                continue

            roznica = abs(saldo_zois - wyciag.saldo_koncowe)
            okres_info = ""
            ostrz_okresu = ""
            if wyciag.miesiac_ostatniej_operacji:
                okres_info = f" | Ostatnia operacja: {wyciag.okres_opisowy}"
                if (wyciag.miesiac_ostatniej_operacji != 12 or
                    (wyciag.rok_ostatniej_operacji and
                     wyciag.rok_ostatniej_operacji != rok_obrachunkowy)):
                    ostrz_okresu = (
                        f" UWAGA: Rachunek zamknięty operacjami z "
                        f"{wyciag.okres_opisowy}, nie z grudnia {rok_obrachunkowy}."
                    )
            bank_str = f" ({wyciag.bank_nazwa})" if wyciag.bank_nazwa else ""

            if roznica < Decimal("0.01"):
                status = StatusAudytu.OSTRZEZ if ostrz_okresu else StatusAudytu.OK
                self._wyniki.append(PunktKontroli(
                    konto=numer_ks, punkt=f"Rachunek {opis}{bank_str}",
                    status=status,
                    uwagi=f"Saldo ZOiS zgodne z wyciągiem.{ostrz_okresu}{okres_info}",
                    wartosc=f"{saldo_zois:,.2f} zł",
                ))
            else:
                self._wyniki.append(PunktKontroli(
                    konto=numer_ks, punkt=f"Rachunek {opis}{bank_str}",
                    status=StatusAudytu.BLAD,
                    uwagi=(f"NIEZGODNOŚĆ: ZOiS={saldo_zois:,.2f} zł | "
                           f"Wyciąg={wyciag.saldo_koncowe:,.2f} zł | "
                           f"Różnica={roznica:,.2f} zł.{ostrz_okresu}{okres_info}"),
                ))

        if rachunki_bez_wyciagu:
            lista = "; ".join(f"{n} ({o}): {s:,.2f} zł" for n, o, s in rachunki_bez_wyciagu)
            self._wyniki.append(PunktKontroli(
                konto="130 (podsum.)",
                punkt=f"Rachunki bez wgranego wyciągu ({len(rachunki_bez_wyciagu)} szt.)",
                status=StatusAudytu.OSTRZEZ,
                uwagi=f"Do uzupełnienia: {lista}",
            ))

    # ─── WERYFIKACJA POSZCZEGÓLNYCH KONT (jak w v2.0) ─────────────────────────

    def _saldo(self, dane_zois, konto):
        return dane_zois.konta.get(konto)

    def _znajdz_syntetyki_po_opisie(
        self,
        dane_zois: "DaneZOiS",
        slowa_kluczowe: List[str],
        grupy_syntetyk: Optional[List[str]] = None,
    ) -> List[Tuple[str, str, Decimal, Decimal]]:
        """
        Wyszukuje syntetyki których opis zawiera któreś ze słów kluczowych.

        Parametry:
          - slowa_kluczowe: np. ["odbiorc", "klient"]
          - grupy_syntetyk: opcjonalnie zawężenie do grup ["20", "22"]

        Zwraca listę: (numer_konta, opis, saldo_Wn, saldo_Ma).
        """
        wynik = []
        slowa_lower = [s.lower() for s in slowa_kluczowe]
        for numer, (wn, ma) in dane_zois.konta.items():
            # Tylko czyste syntetyki (bez myślnika)
            if "-" in numer:
                continue
            if grupy_syntetyk and not any(numer.startswith(g) for g in grupy_syntetyk):
                continue
            opis = dane_zois.opisy.get(numer, "")
            opis_lower = opis.lower()
            if any(s in opis_lower for s in slowa_lower):
                wynik.append((numer, opis, wn, ma))
        return wynik

    def _znajdz_analityki_po_opisie(
        self,
        dane_zois: "DaneZOiS",
        slowa_kluczowe: List[str],
        grupy_syntetyk: Optional[List[str]] = None,
    ) -> List[Tuple[str, str, Decimal, Decimal]]:
        """
        Analogicznie do syntetyk, ale szuka w analitykach (konta z myślnikiem).
        Przykład użycia: szukanie analityki ZUS "220-3" po opisie "ZUS".
        """
        wynik = []
        slowa_lower = [s.lower() for s in slowa_kluczowe]
        for numer, (wn, ma) in dane_zois.konta_analityki.items():
            if "-" not in numer:
                continue
            numer_syn = normalize_konto(numer)
            if grupy_syntetyk and not any(numer_syn.startswith(g) for g in grupy_syntetyk):
                continue
            opis = dane_zois.opisy.get(numer, "")
            opis_lower = opis.lower()
            if any(s in opis_lower for s in slowa_lower):
                wynik.append((numer, opis, wn, ma))
        return wynik

    def _weryfikuj_konto_145(self, dane_zois):
        """
        Weryfikacja środków pieniężnych w drodze. W różnych planach kont:
          - 145 (standard)
          - 133/134 (IGRAPES – "środki w drodze", "rachunek VAT")
        Saldo powinno wynosić 0 na koniec roku.
        """
        # Szukamy po opisie "w drodze" albo "drodze" w grupach 13x/14x
        kandydaci = self._znajdz_syntetyki_po_opisie(
            dane_zois, ["w drodze", "drodze"], grupy_syntetyk=["13", "14"]
        )
        # Jeśli nie znaleziono po opisie - spróbuj klasyczne 145
        if not kandydaci:
            s = self._saldo(dane_zois, "145")
            if s is None:
                self._wyniki.append(PunktKontroli(
                    konto="145", punkt="Środki pieniężne w drodze = 0",
                    status=StatusAudytu.INFO,
                    uwagi="Konto środków w drodze nie wystąpiło.",
                )); return
            kandydaci = [("145", dane_zois.opisy.get("145", "Środki w drodze"),
                         s[0], s[1])]

        # Sprawdzamy każde konto z osobna
        for numer, opis, wn, ma in kandydaci:
            if wn == Decimal("0") and ma == Decimal("0"):
                self._wyniki.append(PunktKontroli(
                    konto=numer, punkt=f"{opis} = 0",
                    status=StatusAudytu.OK, uwagi="Saldo wynosi 0.",
                ))
            else:
                self._wyniki.append(PunktKontroli(
                    konto=numer, punkt=f"{opis} = 0",
                    status=StatusAudytu.BLAD,
                    uwagi=f"Saldo ≠ 0! Wn={wn:,.2f}, Ma={ma:,.2f} zł.",
                ))

    def _weryfikuj_konto_200(self, dane_zois):
        """
        Rozrachunki z odbiorcami – grupa 20x (np. 200, 201).
        Wyszukujemy po opisie zawierającym "odbiorc" lub "klient".
        Oczekiwane saldo: Wn (należności).
        """
        kandydaci = self._znajdz_syntetyki_po_opisie(
            dane_zois, ["odbiorc", "klient"], grupy_syntetyk=["20"]
        )
        if not kandydaci:
            # Fallback: klasyczne 200
            s = self._saldo(dane_zois, "200")
            if s is None:
                self._wyniki.append(PunktKontroli(
                    konto="20x", punkt="Rozrachunki z odbiorcami",
                    status=StatusAudytu.BRAK,
                    uwagi="Brak konta rozrachunków z odbiorcami (20x).",
                )); return
            kandydaci = [("200", dane_zois.opisy.get("200", "Rozrachunki - odbiorcy"),
                         s[0], s[1])]

        for numer, opis, wn, ma in kandydaci:
            if ma > Decimal("0"):
                self._wyniki.append(PunktKontroli(
                    konto=numer, punkt=f"{opis} – strona salda",
                    status=StatusAudytu.BLAD,
                    uwagi=f"Saldo Ma={ma:,.2f} zł – BŁĄD. Brak faktury sprzedaży lub nadpłata.",
                ))
            else:
                self._wyniki.append(PunktKontroli(
                    konto=numer, punkt=f"{opis} – strona salda",
                    status=StatusAudytu.OK,
                    uwagi=f"Saldo Wn={wn:,.2f} zł – należności.",
                ))

    def _weryfikuj_konto_202(self, dane_zois):
        """
        Rozrachunki z dostawcami – grupa 20x (np. 202, 210, 211).
        Wyszukujemy po opisie zawierającym "dostawc".
        Oczekiwane saldo: Ma (zobowiązania).
        """
        kandydaci = self._znajdz_syntetyki_po_opisie(
            dane_zois, ["dostawc"], grupy_syntetyk=["20", "21"]
        )
        if not kandydaci:
            s = self._saldo(dane_zois, "202")
            if s is None:
                self._wyniki.append(PunktKontroli(
                    konto="20x/21x", punkt="Rozrachunki z dostawcami",
                    status=StatusAudytu.BRAK,
                    uwagi="Brak konta rozrachunków z dostawcami.",
                )); return
            kandydaci = [("202", dane_zois.opisy.get("202", "Rozrachunki - dostawcy"),
                         s[0], s[1])]

        for numer, opis, wn, ma in kandydaci:
            if wn > Decimal("0"):
                self._wyniki.append(PunktKontroli(
                    konto=numer, punkt=f"{opis} – strona salda",
                    status=StatusAudytu.BLAD,
                    uwagi=f"Saldo Wn={wn:,.2f} zł – BŁĄD. Nadpłata lub brak faktury zakupu.",
                ))
            else:
                self._wyniki.append(PunktKontroli(
                    konto=numer, punkt=f"{opis} – strona salda",
                    status=StatusAudytu.OK,
                    uwagi=f"Saldo Ma={ma:,.2f} zł – zobowiązania.",
                ))

    def _weryfikuj_konto_230(self, dane_zois):
        s = self._saldo(dane_zois, "230")
        if s is None:
            self._wyniki.append(PunktKontroli(
                konto="230", punkt="Wynagrodzenia",
                status=StatusAudytu.INFO, uwagi="Konto 230 nie wystąpiło.",
            )); return
        wn, ma = s
        if wn > Decimal("0"):
            self._wyniki.append(PunktKontroli(
                konto="230", punkt="Rozrachunki z pracownikami",
                status=StatusAudytu.OSTRZEZ,
                uwagi=f"Saldo Wn={wn:,.2f} zł – brak LP lub nadpłata.",
            ))
        elif ma > Decimal("0"):
            self._wyniki.append(PunktKontroli(
                konto="230", punkt="Rozrachunki z pracownikami",
                status=StatusAudytu.OK,
                uwagi=f"Saldo Ma={ma:,.2f} zł – niezapłacone (dopuszczalne).",
            ))
        else:
            self._wyniki.append(PunktKontroli(
                konto="230", punkt="Rozrachunki z pracownikami",
                status=StatusAudytu.OK, uwagi="Saldo 230 = 0.",
            ))

    def _weryfikuj_konto_229(self, dane_zois):
        """
        Zobowiązania ZUS. W różnych planach kont:
          - 229 (klasyczny standard)
          - 220-3 w IGRAPES (analityka pod 220 Rozrachunki z budżetami)
          - 221-x w innych
        Szuka po opisie "zus" lub "ubezpiecz" w grupach 22x.
        Oczekiwane saldo: Ma (zobowiązanie).
        """
        # Najpierw szukaj w syntetykach
        kandydaci = self._znajdz_syntetyki_po_opisie(
            dane_zois, ["zus"], grupy_syntetyk=["22"]
        )
        # Potem w analitykach (IGRAPES: 220-3 ZUS)
        kandydaci_ana = self._znajdz_analityki_po_opisie(
            dane_zois, ["zus"], grupy_syntetyk=["22"]
        )
        # Gdy mamy analityki, eliminujemy duplikat - syntetykę tej samej grupy
        if kandydaci_ana:
            syntetyki_do_pominiecia = {
                normalize_konto(n) for n, _, _, _ in kandydaci_ana
            }
            kandydaci = [
                k for k in kandydaci if k[0] not in syntetyki_do_pominiecia
            ]
        kandydaci.extend(kandydaci_ana)

        if not kandydaci:
            # Fallback: klasyczne 229
            s = self._saldo(dane_zois, "229")
            if s is None:
                self._wyniki.append(PunktKontroli(
                    konto="ZUS", punkt="Zobowiązanie ZUS",
                    status=StatusAudytu.BRAK,
                    uwagi="Brak konta ZUS (229/220-x/ubezpieczeń).",
                )); return
            kandydaci = [("229", dane_zois.opisy.get("229", "ZUS"), s[0], s[1])]

        for numer, opis, wn, ma in kandydaci:
            if wn > Decimal("0"):
                self._wyniki.append(PunktKontroli(
                    konto=numer, punkt=f"{opis} – zobowiązanie ZUS",
                    status=StatusAudytu.OSTRZEZ,
                    uwagi=f"Saldo Wn={wn:,.2f} zł – nadpłata lub błąd DRA.",
                ))
            elif ma > Decimal("0"):
                self._wyniki.append(PunktKontroli(
                    konto=numer, punkt=f"{opis} – zobowiązanie ZUS",
                    status=StatusAudytu.OK,
                    uwagi=f"Saldo Ma={ma:,.2f} zł. Porównaj z DRA.",
                ))
            else:
                self._wyniki.append(PunktKontroli(
                    konto=numer, punkt=f"{opis} – zobowiązanie ZUS",
                    status=StatusAudytu.OK, uwagi="Saldo wynosi 0.",
                ))

    def _weryfikuj_konto_220(self, dane_zois):
        """
        Zobowiązania podatkowe – PIT-4 (podatek od płac) oraz podatek dochodowy.
        W różnych planach kont:
          - 220 (klasyczny - cały US)
          - 220-1 Podatek dochodowy, 220-2 Podatek od płac (IGRAPES)
          - 222-x (inne systemy)
        Szuka po opisie "podatek" w grupie 22x (bez ZUS).
        Oczekiwane saldo: Ma (zobowiązanie podatkowe).
        """
        kandydaci = self._znajdz_syntetyki_po_opisie(
            dane_zois, ["podatek", "budżet", "urz"], grupy_syntetyk=["22"]
        )
        kandydaci_ana = self._znajdz_analityki_po_opisie(
            dane_zois, ["podatek", "pit"], grupy_syntetyk=["22"]
        )
        # Wyłącz analityki ZUS (już obsłużone w _weryfikuj_konto_229)
        kandydaci_ana = [
            k for k in kandydaci_ana if "zus" not in k[1].lower()
        ]
        # Gdy mamy analityki, pomijamy syntetyki tej samej grupy (duplikacja)
        if kandydaci_ana:
            syntetyki_do_pominiecia = {
                normalize_konto(n) for n, _, _, _ in kandydaci_ana
            }
            kandydaci = [
                k for k in kandydaci if k[0] not in syntetyki_do_pominiecia
            ]
        kandydaci.extend(kandydaci_ana)

        if not kandydaci:
            s = self._saldo(dane_zois, "220")
            if s is None:
                self._wyniki.append(PunktKontroli(
                    konto="US/PIT", punkt="Zobowiązanie podatkowe",
                    status=StatusAudytu.BRAK,
                    uwagi="Brak konta podatkowego (220/22x).",
                )); return
            kandydaci = [("220", dane_zois.opisy.get("220", "US"), s[0], s[1])]

        for numer, opis, wn, ma in kandydaci:
            if wn > Decimal("0"):
                self._wyniki.append(PunktKontroli(
                    konto=numer, punkt=f"{opis} – zobowiązanie podatkowe",
                    status=StatusAudytu.OSTRZEZ,
                    uwagi=f"Saldo Wn={wn:,.2f} zł – nadpłata lub brak dekretacji XII.",
                ))
            elif ma > Decimal("0"):
                self._wyniki.append(PunktKontroli(
                    konto=numer, punkt=f"{opis} – zobowiązanie podatkowe",
                    status=StatusAudytu.OK,
                    uwagi=f"Saldo Ma={ma:,.2f} zł. Porównaj z deklaracją PIT-4R/8AR.",
                ))
            else:
                self._wyniki.append(PunktKontroli(
                    konto=numer, punkt=f"{opis} – zobowiązanie podatkowe",
                    status=StatusAudytu.OK, uwagi="Saldo wynosi 0.",
                ))

    def _weryfikuj_konto_700(self, dane_zois):
        """
        Weryfikuje konta grupy 70 (przychody ze sprzedaży: 700-709).
        W praktyce podmioty używają różnych numerów: 700 (Sprzedaż), 701, 702
        (Sprzedaż usług), 703 (Sprzedaż towarów) itp.
        """
        # Sumuj wszystkie konta z grupy 70x (700-709)
        konta_70 = {
            k: v for k, v in dane_zois.konta.items()
            if k.isdigit() and len(k) == 3 and k.startswith("70")
        }
        if not konta_70:
            self._wyniki.append(PunktKontroli(
                konto="70x", punkt="Przychody ze sprzedaży",
                status=StatusAudytu.BRAK,
                uwagi="CRITICAL: Brak kont grupy 70x (700-709).",
            )); return

        suma_wn = sum((wn for wn, _ in konta_70.values()), Decimal("0"))
        suma_ma = sum((ma for _, ma in konta_70.values()), Decimal("0"))
        konta_str = ", ".join(sorted(konta_70.keys()))

        if suma_wn > Decimal("0"):
            bledne = [k for k, (wn, _) in konta_70.items() if wn > Decimal("0")]
            self._wyniki.append(PunktKontroli(
                konto="70x", punkt="Przychody ze sprzedaży – strona salda",
                status=StatusAudytu.BLAD,
                uwagi=f"Saldo Wn={suma_wn:,.2f} zł – BŁĄD! "
                      f"Konta wynikowe Ma. Błędne konta: {', '.join(bledne)}.",
            ))
        elif suma_ma > Decimal("0"):
            self._wyniki.append(PunktKontroli(
                konto="70x", punkt=f"Przychody ze sprzedaży ({konta_str})",
                status=StatusAudytu.OK,
                uwagi=f"Saldo Ma={suma_ma:,.2f} zł – poprawne.",
                wartosc=f"{suma_ma:,.2f} zł",
            ))
        else:
            self._wyniki.append(PunktKontroli(
                konto="70x", punkt="Przychody ze sprzedaży",
                status=StatusAudytu.OSTRZEZ,
                uwagi="Zerowe salda – brak sprzedaży w okresie?",
            ))

    def _weryfikuj_grupe_4(self, dane_zois):
        konta = {k: v for k, v in dane_zois.konta.items() if get_grupa(k) == 4}
        if not konta:
            self._wyniki.append(PunktKontroli(
                konto="Grupa 4", punkt="Koszty rodzajowe (400–499)",
                status=StatusAudytu.BRAK, uwagi="Brak kont grupy 4.",
            )); return
        bledne = []
        suma = Decimal("0")
        for k, (wn, ma) in konta.items():
            suma += wn
            if ma > Decimal("0"):
                opis = dane_zois.opisy.get(k, "")
                bledne.append(f"{k} ({opis}): Ma={ma:,.2f} zł")
        if bledne:
            self._wyniki.append(PunktKontroli(
                konto="Grupa 4", punkt="Koszty rodzajowe – tylko Saldo Wn",
                status=StatusAudytu.BLAD,
                uwagi=(f"Konta z Saldem Ma ({len(bledne)} szt.): "
                       + "; ".join(bledne[:5])
                       + (" ..." if len(bledne) > 5 else "")),
            ))
        else:
            self._wyniki.append(PunktKontroli(
                konto="Grupa 4", punkt="Koszty rodzajowe – tylko Saldo Wn",
                status=StatusAudytu.OK,
                uwagi=f"Wszystkie {len(konta)} kont poprawne. Suma: {suma:,.2f} zł.",
            ))

    # ─── WERYFIKACJA BILANSU (v3.0 – rozszerzona) ────────────────────────────

    def _weryfikuj_bilans(self, dane_bilans: Optional[DaneBilansu]):
        if dane_bilans is None:
            self._wyniki.append(PunktKontroli(
                konto="Bilans", punkt="Suma bilansowa",
                status=StatusAudytu.BRAK, uwagi="Nie dostarczono pliku Bilansu.",
            )); return

        TOL = Decimal("0.01")

        # 1. Spójność wewnętrzna aktywów: A + B + C + D = Suma aktywów
        suma_obl_a = (
            dane_bilans.aktywa_trwale_biezacy + dane_bilans.aktywa_obrotowe_biezacy
            + dane_bilans.nalezne_wplaty_biezacy + dane_bilans.udzialy_wlasne_biezacy
        )
        if dane_bilans.suma_aktywow_biezacy > Decimal("0"):
            roznica = abs(suma_obl_a - dane_bilans.suma_aktywow_biezacy)
            if roznica <= TOL:
                self._wyniki.append(PunktKontroli(
                    konto="Bilans", punkt="Spójność wewnętrzna aktywów (A+B+C+D=Suma)",
                    status=StatusAudytu.OK,
                    uwagi=(f"Aktywa trwałe+obrotowe+C+D = Suma aktywów "
                           f"({dane_bilans.suma_aktywow_biezacy:,.2f} zł)."),
                ))
            else:
                self._wyniki.append(PunktKontroli(
                    konto="Bilans", punkt="Spójność wewnętrzna aktywów (A+B+C+D=Suma)",
                    status=StatusAudytu.BLAD,
                    uwagi=(f"NIEZGODNOŚĆ! Obliczona suma={suma_obl_a:,.2f}, "
                           f"Suma w bilansie={dane_bilans.suma_aktywow_biezacy:,.2f} | "
                           f"Różnica={roznica:,.2f} zł."),
                ))

        # 2. Spójność wewnętrzna pasywów: A + B = Suma pasywów
        suma_obl_p = dane_bilans.kapital_wlasny_biezacy + dane_bilans.zobowiazania_biezacy
        if dane_bilans.suma_pasywow_biezacy > Decimal("0"):
            roznica = abs(suma_obl_p - dane_bilans.suma_pasywow_biezacy)
            if roznica <= TOL:
                self._wyniki.append(PunktKontroli(
                    konto="Bilans", punkt="Spójność wewnętrzna pasywów (A+B=Suma)",
                    status=StatusAudytu.OK,
                    uwagi=(f"Kapitał własny+Zobowiązania = Suma pasywów "
                           f"({dane_bilans.suma_pasywow_biezacy:,.2f} zł)."),
                ))
            else:
                self._wyniki.append(PunktKontroli(
                    konto="Bilans", punkt="Spójność wewnętrzna pasywów (A+B=Suma)",
                    status=StatusAudytu.BLAD,
                    uwagi=(f"NIEZGODNOŚĆ! Obliczona suma={suma_obl_p:,.2f}, "
                           f"Suma w bilansie={dane_bilans.suma_pasywow_biezacy:,.2f} | "
                           f"Różnica={roznica:,.2f} zł."),
                ))

        # 3. Aktywa = Pasywa (rok bieżący)
        rb = abs(dane_bilans.suma_aktywow_biezacy - dane_bilans.suma_pasywow_biezacy)
        if rb <= TOL:
            self._wyniki.append(PunktKontroli(
                konto="Bilans", punkt="Suma bilansowa rok bieżący",
                status=StatusAudytu.OK,
                uwagi=f"Aktywa = Pasywa = {dane_bilans.suma_aktywow_biezacy:,.2f} zł.",
            ))
        else:
            self._wyniki.append(PunktKontroli(
                konto="Bilans", punkt="Suma bilansowa rok bieżący",
                status=StatusAudytu.BLAD,
                uwagi=(f"NIEZGODNOŚĆ! Aktywa={dane_bilans.suma_aktywow_biezacy:,.2f} ≠ "
                       f"Pasywa={dane_bilans.suma_pasywow_biezacy:,.2f} | Δ={rb:,.2f} zł."),
            ))

        # 4. Aktywa = Pasywa (rok ubiegły)
        if dane_bilans.suma_aktywow_ubiegly > Decimal("0") or dane_bilans.suma_pasywow_ubiegly > Decimal("0"):
            ru = abs(dane_bilans.suma_aktywow_ubiegly - dane_bilans.suma_pasywow_ubiegly)
            if ru <= TOL:
                self._wyniki.append(PunktKontroli(
                    konto="Bilans", punkt="Suma bilansowa rok ubiegły",
                    status=StatusAudytu.OK,
                    uwagi=f"Dane porównawcze: {dane_bilans.suma_aktywow_ubiegly:,.2f} zł.",
                ))
            else:
                self._wyniki.append(PunktKontroli(
                    konto="Bilans", punkt="Suma bilansowa rok ubiegły",
                    status=StatusAudytu.BLAD,
                    uwagi=(f"NIEZGODNOŚĆ! Aktywa={dane_bilans.suma_aktywow_ubiegly:,.2f} ≠ "
                           f"Pasywa={dane_bilans.suma_pasywow_ubiegly:,.2f} | Δ={ru:,.2f} zł."),
                ))

    # ─── WERYFIKACJA RZiS (v3.0 – nowa) ──────────────────────────────────────

    def _weryfikuj_rzis(self, dane_rzis: Optional[DaneRZiS]):
        if dane_rzis is None:
            self._wyniki.append(PunktKontroli(
                konto="RZiS", punkt="Rachunek Zysków i Strat",
                status=StatusAudytu.BRAK, uwagi="Nie dostarczono pliku RZiS.",
            )); return

        TOL = Decimal("1.00")  # tolerancja 1 zł (zaokrąglenia Symfonii)

        # Sprawdzenia spójności matematycznej RZiS
        reguly = [
            ("C = A − B (Zysk ze sprzedaży)",
             dane_rzis.przychody_sprzedazy[0] - dane_rzis.koszty_operacyjne[0],
             dane_rzis.zysk_sprzedazy[0]),
            ("F = C + D − E (Wynik operacyjny)",
             dane_rzis.zysk_sprzedazy[0]
              + dane_rzis.pozostale_przych_oper[0]
              - dane_rzis.pozostale_koszty_oper[0],
             dane_rzis.zysk_dzialalnosci_oper[0]),
            ("I = F + G − H (Zysk brutto)",
             dane_rzis.zysk_dzialalnosci_oper[0]
              + dane_rzis.przychody_finansowe[0]
              - dane_rzis.koszty_finansowe[0],
             dane_rzis.zysk_brutto[0]),
            ("L = I − J − K (Zysk netto)",
             dane_rzis.zysk_brutto[0]
              - dane_rzis.podatek_dochodowy[0]
              - dane_rzis.pozostale_zmniejszenia[0],
             dane_rzis.zysk_netto[0]),
        ]

        for nazwa, obliczona, podana in reguly:
            roznica = abs(obliczona - podana)
            if roznica <= TOL:
                self._wyniki.append(PunktKontroli(
                    konto="RZiS", punkt=nazwa,
                    status=StatusAudytu.OK,
                    uwagi=f"Spójność matematyczna OK ({podana:,.2f} zł).",
                ))
            else:
                self._wyniki.append(PunktKontroli(
                    konto="RZiS", punkt=nazwa,
                    status=StatusAudytu.BLAD,
                    uwagi=(f"NIESPÓJNOŚĆ! Obliczona={obliczona:,.2f}, "
                           f"Podana={podana:,.2f} | Różnica={roznica:,.2f} zł."),
                ))

        # Informacyjnie – zysk/strata netto
        zn = dane_rzis.zysk_netto[0]
        self._wyniki.append(PunktKontroli(
            konto="RZiS", punkt="Zysk (strata) netto",
            status=StatusAudytu.INFO,
            uwagi=f"Wynik finansowy roku bieżącego: {zn:,.2f} zł.",
            wartosc=f"{zn:,.2f} zł",
        ))

    # ─── REGUŁY KRZYŻOWE (v3.0 – nowa) ───────────────────────────────────────

    def _weryfikuj_krzyzowe(
        self,
        dane_zois: Optional[DaneZOiS],
        dane_bilans: Optional[DaneBilansu],
        dane_rzis: Optional[DaneRZiS],
    ):
        """Weryfikacja krzyżowa między ZOiS, Bilansem i RZiS."""
        TOL = Decimal("1.00")

        # 1. Wynik netto RZiS = Wynik netto w Bilansie (pasywa A.VI)
        if dane_rzis is not None and dane_bilans is not None:
            zn_rzis = dane_rzis.zysk_netto[0]
            zn_bil = dane_bilans.wynik_netto_biezacy
            if zn_rzis == Decimal("0") and zn_bil == Decimal("0"):
                pass  # nie mamy danych
            else:
                roznica = abs(zn_rzis - zn_bil)
                if roznica <= TOL:
                    self._wyniki.append(PunktKontroli(
                        konto="RZiS↔Bilans",
                        punkt="Wynik netto: RZiS (L) = Bilans (Pasywa A.VI)",
                        status=StatusAudytu.OK,
                        uwagi=f"Wynik netto zgodny: {zn_rzis:,.2f} zł.",
                    ))
                else:
                    self._wyniki.append(PunktKontroli(
                        konto="RZiS↔Bilans",
                        punkt="Wynik netto: RZiS (L) = Bilans (Pasywa A.VI)",
                        status=StatusAudytu.BLAD,
                        uwagi=(f"NIEZGODNOŚĆ! RZiS={zn_rzis:,.2f} zł, "
                               f"Bilans={zn_bil:,.2f} zł | Δ={roznica:,.2f} zł."),
                    ))

        # 2. Grupa 70x (Ma) ≈ RZiS poz. A (Przychody ze sprzedaży)
        if dane_zois is not None and dane_rzis is not None:
            konta_70 = {
                k: v for k, v in dane_zois.konta.items()
                if k.isdigit() and len(k) == 3 and k.startswith("70")
            }
            if konta_70:
                suma_70_ma = sum((ma for _, ma in konta_70.values()), Decimal("0"))
                przych = dane_rzis.przychody_sprzedazy[0]
                konta_str = "+".join(sorted(konta_70.keys()))
                if przych > Decimal("0") and suma_70_ma > Decimal("0"):
                    roznica = abs(suma_70_ma - przych)
                    if roznica <= TOL:
                        self._wyniki.append(PunktKontroli(
                            konto="ZOiS↔RZiS",
                            punkt=f"Grupa 70x (Ma) ≈ RZiS poz. A (Przychody)",
                            status=StatusAudytu.OK,
                            uwagi=f"Przychody zgodne ({konta_str}): {suma_70_ma:,.2f} zł.",
                        ))
                    else:
                        self._wyniki.append(PunktKontroli(
                            konto="ZOiS↔RZiS",
                            punkt=f"Grupa 70x (Ma) ≈ RZiS poz. A (Przychody)",
                            status=StatusAudytu.OSTRZEZ,
                            uwagi=(f"Różnica: ZOiS {konta_str}={suma_70_ma:,.2f} zł, "
                                   f"RZiS A={przych:,.2f} zł | Δ={roznica:,.2f} zł."),
                        ))

        # 3. Suma grupy 4 (Wn) ≈ RZiS poz. B (Koszty operacyjne)
        if dane_zois is not None and dane_rzis is not None:
            grupa_4_wn = sum(
                (wn for k, (wn, _) in dane_zois.konta.items() if get_grupa(k) == 4),
                Decimal("0")
            )
            koszty_b = dane_rzis.koszty_operacyjne[0]
            if grupa_4_wn > Decimal("0") and koszty_b > Decimal("0"):
                roznica = abs(grupa_4_wn - koszty_b)
                if roznica <= TOL:
                    self._wyniki.append(PunktKontroli(
                        konto="ZOiS↔RZiS",
                        punkt="Suma grupy 4 (Wn) ≈ RZiS poz. B (Koszty)",
                        status=StatusAudytu.OK,
                        uwagi=f"Koszty operacyjne zgodne: {koszty_b:,.2f} zł.",
                    ))
                else:
                    self._wyniki.append(PunktKontroli(
                        konto="ZOiS↔RZiS",
                        punkt="Suma grupy 4 (Wn) ≈ RZiS poz. B (Koszty)",
                        status=StatusAudytu.OSTRZEZ,
                        uwagi=(f"Różnica: ZOiS grupa 4={grupa_4_wn:,.2f} zł, "
                               f"RZiS B={koszty_b:,.2f} zł | Δ={roznica:,.2f} zł. "
                               "Sprawdź amortyzację i pozostałe pozycje."),
                    ))

    # ─── R2: WERYFIKACJA OBROTÓW Wn=Ma (v3.2) ────────────────────────────────

    def _weryfikuj_obroty_wn_ma(self, dane_zois: DaneZOiS):
        """
        Weryfikuje fundamentalną zasadę podwójnego zapisu – suma obrotów Wn
        musi równać się sumie obrotów Ma dla 3 par kolumn:
          - BO (Bilans Otwarcia) – błąd tu oznacza błędny bilans zamknięcia
            poprzedniego roku
          - Obroty miesiąca – błąd w dekretacjach bieżącego miesiąca
          - Narastająco – suma wszystkich dekretacji roku
        """
        s = dane_zois.sumy_razem
        if not s.wykryto:
            self._wyniki.append(PunktKontroli(
                konto="ZOiS (Wn=Ma)",
                punkt="Kontrola dwustronnego zapisu",
                status=StatusAudytu.BRAK,
                uwagi="Nie znaleziono wiersza 'Suma razem' w ZOiS. "
                      "Starszy format eksportu lub XLSX bez wiersza sumarycznego.",
            ))
            return

        TOL = Decimal("0.01")
        reguly = [
            ("BO (Bilans Otwarcia)", s.bo_wn, s.bo_ma),
            ("Obroty bieżącego miesiąca", s.obroty_wn, s.obroty_ma),
            ("Obroty narastająco (cały rok)", s.narastajaco_wn, s.narastajaco_ma),
        ]

        for nazwa, wn, ma in reguly:
            roznica = abs(wn - ma)
            if roznica <= TOL:
                self._wyniki.append(PunktKontroli(
                    konto="ZOiS (Wn=Ma)",
                    punkt=f"Dwustronny zapis: {nazwa}",
                    status=StatusAudytu.OK,
                    uwagi=f"Wn = Ma = {wn:,.2f} zł ✓",
                ))
            else:
                self._wyniki.append(PunktKontroli(
                    konto="ZOiS (Wn=Ma)",
                    punkt=f"Dwustronny zapis: {nazwa}",
                    status=StatusAudytu.BLAD,
                    uwagi=(f"NIEZBILANSOWANE! Wn={wn:,.2f} zł ≠ Ma={ma:,.2f} zł | "
                           f"Różnica={roznica:,.2f} zł. CRITICAL: jednostronna dekretacja."),
                ))

    # ─── R1: WERYFIKACJA KONTA 860 (v3.2) ────────────────────────────────────

    def _weryfikuj_konto_860(
        self,
        dane_zois: DaneZOiS,
        dane_rzis: Optional[DaneRZiS],
    ):
        """
        Automatyczne wykrywanie stanu klienta (przed/po zamknięciu roku) i
        stosowanie odpowiedniej reguły weryfikacji wyniku finansowego na 860:

          Stan "Przed zamknięciem": saldo 860 ≠ 0 → musi = L z RZiS
          Stan "Po zamknięciu":     saldo 860 = 0, obroty Ma − Wn = L z RZiS

        Dodatkowe ostrzeżenie: jeśli 860 ma zero obrotów narastająco a grupy
        4 i 7 mają salda – klient nie wykonał przeksięgowań zamykających.
        """
        saldo_860 = dane_zois.konta.get("860")
        obroty_860 = dane_zois.obroty_narastajaco.get("860")

        if saldo_860 is None and obroty_860 is None:
            self._wyniki.append(PunktKontroli(
                konto="860", punkt="Konto 860 (Wynik finansowy)",
                status=StatusAudytu.BRAK,
                uwagi="Konto 860 nie wystąpiło w ZOiS – brak wyniku finansowego.",
            ))
            return

        saldo_wn, saldo_ma = saldo_860 or (Decimal("0"), Decimal("0"))
        obr_wn, obr_ma = obroty_860 or (Decimal("0"), Decimal("0"))
        saldo_netto = saldo_ma - saldo_wn  # dodatni = Ma (zysk)

        TOL = Decimal("1.00")

        # ── Detekcja stanu klienta ─────────────────────────────────────────
        ma_obroty = (obr_wn > Decimal("0") or obr_ma > Decimal("0"))
        ma_saldo  = (saldo_wn > Decimal("0") or saldo_ma > Decimal("0"))

        if ma_obroty and not ma_saldo:
            stan = "po_zamknieciu"
            stan_opis = "po zamknięciu roku (wynik przeksięgowany na 821/863)"
            wynik_z_860 = obr_ma - obr_wn
        elif ma_saldo:
            stan = "przed_zamknieciem"
            stan_opis = "przed zamknięciem roku (wynik na saldzie 860)"
            wynik_z_860 = saldo_netto
        else:
            # Brak obrotów i salda – klient nie zrobił przeksięgowań
            stan = "brak_zamkniecia"
            stan_opis = "brak przeksięgowań"
            wynik_z_860 = Decimal("0")

        self._wyniki.append(PunktKontroli(
            konto="860", punkt="Wykryty stan cyklu zamknięcia",
            status=StatusAudytu.INFO,
            uwagi=f"Wykryto stan: {stan_opis}.",
        ))

        # ── Porównanie z RZiS ──────────────────────────────────────────────
        if dane_rzis is not None and stan != "brak_zamkniecia":
            zysk_rzis = dane_rzis.zysk_netto[0]
            roznica = abs(wynik_z_860 - zysk_rzis)
            if roznica <= TOL:
                self._wyniki.append(PunktKontroli(
                    konto="860↔RZiS",
                    punkt="Wynik netto: konto 860 = RZiS poz. L",
                    status=StatusAudytu.OK,
                    uwagi=f"Wynik spójny: {wynik_z_860:,.2f} zł.",
                ))
            else:
                self._wyniki.append(PunktKontroli(
                    konto="860↔RZiS",
                    punkt="Wynik netto: konto 860 = RZiS poz. L",
                    status=StatusAudytu.BLAD,
                    uwagi=(f"NIEZGODNOŚĆ! 860={wynik_z_860:,.2f} zł, "
                           f"RZiS L={zysk_rzis:,.2f} zł | Δ={roznica:,.2f} zł."),
                ))

        # ── Ostrzeżenie: brak przeksięgowań (grupa 4/7 ma salda, 860 puste) ──
        if stan == "brak_zamkniecia":
            grupa_4 = sum(
                (wn for k, (wn, _) in dane_zois.konta.items() if get_grupa(k) == 4),
                Decimal("0"),
            )
            grupa_7_ma = sum(
                (ma for k, (_, ma) in dane_zois.konta.items() if get_grupa(k) == 7),
                Decimal("0"),
            )
            if grupa_4 > Decimal("0") or grupa_7_ma > Decimal("0"):
                self._wyniki.append(PunktKontroli(
                    konto="860", punkt="Przeksięgowania zamykające",
                    status=StatusAudytu.OSTRZEZ,
                    uwagi=(f"Grupa 4 (koszty)={grupa_4:,.2f} zł, "
                           f"Grupa 7 (przychody)={grupa_7_ma:,.2f} zł, "
                           f"ale konto 860 jest puste. "
                           "Klient nie wykonał przeksięgowań zamykających rok."),
                ))

    # ─── R3: WERYFIKACJA KAPITAŁU vs KRS (v3.2) ──────────────────────────────

    def _weryfikuj_krs(
        self,
        dane_zois:   Optional[DaneZOiS],
        dane_bilans: Optional[DaneBilansu],
        dane_krs:    DaneKRS,
    ):
        """
        Weryfikuje zgodność kapitału zakładowego z rejestrem KRS.

        Priorytetowo porównuje z saldem konta rozpoznanego po opisie
        "kapitał zakładowy/podstawowy" w ZOiS. Fallback: pozycja A.I w pasywach
        Bilansu (kapitał podstawowy).

        Dodatkowo weryfikuje nazwę podmiotu (fuzzy match tolerujący skróty
        jak "SP. Z O.O." vs "SPÓŁKA Z OGRANICZONĄ ODPOWIEDZIALNOŚCIĄ").
        """
        # Jeśli błąd pobrania KRS – zaznacz i zakończ
        if dane_krs.blad:
            self._wyniki.append(PunktKontroli(
                konto="KRS", punkt="Pobranie danych z rejestru KRS",
                status=StatusAudytu.BRAK,
                uwagi=f"Nie udało się pobrać danych: {dane_krs.blad}",
            ))
            return

        # ── Info o pobranych danych ────────────────────────────────────────
        self._wyniki.append(PunktKontroli(
            konto="KRS", punkt=f"Dane z KRS: {dane_krs.nazwa}",
            status=StatusAudytu.INFO,
            uwagi=(f"KRS: {dane_krs.numer_krs} | NIP: {dane_krs.nip} | "
                   f"REGON: {dane_krs.regon} | "
                   f"Kapitał zakładowy: {dane_krs.kapital_zakladowy:,.2f} zł"),
        ))

        # ── Weryfikacja kapitału zakładowego ────────────────────────────────
        if dane_krs.kapital_zakladowy <= Decimal("0"):
            self._wyniki.append(PunktKontroli(
                konto="KRS↔ZOiS", punkt="Kapitał zakładowy",
                status=StatusAudytu.BRAK,
                uwagi="API KRS nie zwróciło wartości kapitału zakładowego.",
            ))
        else:
            TOL = Decimal("0.01")
            porownanie_wykonane = False

            # Priorytetowo: saldo konta "kapitał zakładowy" w ZOiS
            if dane_zois is not None:
                konta_kapital = [
                    k for k, opis in dane_zois.opisy.items()
                    if "-" not in k and (
                        "kapitał zakładowy" in opis.lower()
                        or "kapitał podstawowy" in opis.lower()
                    )
                ]
                if konta_kapital:
                    kapital_zois = Decimal("0")
                    konta_uzyte = []
                    for k in konta_kapital:
                        wn, ma = dane_zois.konta.get(k, (Decimal("0"), Decimal("0")))
                        kapital_zois += (ma - wn)  # kapitał zawsze Ma (pasywa)
                        konta_uzyte.append(k)

                    roznica = abs(kapital_zois - dane_krs.kapital_zakladowy)
                    if roznica <= TOL:
                        self._wyniki.append(PunktKontroli(
                            konto="KRS↔ZOiS",
                            punkt=f"Kapitał zakładowy: ZOiS ({', '.join(konta_uzyte)}) = KRS",
                            status=StatusAudytu.OK,
                            uwagi=f"Kapitał zgodny: {kapital_zois:,.2f} zł.",
                        ))
                    else:
                        self._wyniki.append(PunktKontroli(
                            konto="KRS↔ZOiS",
                            punkt=f"Kapitał zakładowy: ZOiS ({', '.join(konta_uzyte)}) = KRS",
                            status=StatusAudytu.BLAD,
                            uwagi=(f"NIEZGODNOŚĆ! ZOiS={kapital_zois:,.2f} zł, "
                                   f"KRS={dane_krs.kapital_zakladowy:,.2f} zł | "
                                   f"Δ={roznica:,.2f} zł. "
                                   "Sprawdź uchwały o zmianie kapitału."),
                        ))
                    porownanie_wykonane = True

            # Fallback: pozycja A.I w pasywach Bilansu (jeśli dostępny)
            if not porownanie_wykonane and dane_bilans is not None:
                # UWAGA: DaneBilansu v3.1 nie rozbija pasywów na A.I – ma tylko
                # kapital_wlasny_biezacy (cała pozycja A). Dlatego ten fallback
                # daje tylko OSTRZEŻENIE a nie porównanie co do grosza.
                kap_wlasny = dane_bilans.kapital_wlasny_biezacy
                if kap_wlasny >= dane_krs.kapital_zakladowy:
                    self._wyniki.append(PunktKontroli(
                        konto="KRS↔Bilans",
                        punkt="Kapitał zakładowy w Bilansie",
                        status=StatusAudytu.INFO,
                        uwagi=(f"Brak konta 'kapitał zakładowy' w ZOiS. "
                               f"Kapitał własny w Bilansie ({kap_wlasny:,.2f} zł) "
                               f"zawiera KRS ({dane_krs.kapital_zakladowy:,.2f} zł) "
                               "– szczegółowa weryfikacja niemożliwa bez rozbicia "
                               "pozycji A.I."),
                    ))
                else:
                    self._wyniki.append(PunktKontroli(
                        konto="KRS↔Bilans",
                        punkt="Kapitał zakładowy w Bilansie",
                        status=StatusAudytu.BLAD,
                        uwagi=(f"Kapitał własny w Bilansie ({kap_wlasny:,.2f} zł) "
                               f"mniejszy niż kapitał zakładowy z KRS "
                               f"({dane_krs.kapital_zakladowy:,.2f} zł)."),
                    ))
                porownanie_wykonane = True

            if not porownanie_wykonane:
                self._wyniki.append(PunktKontroli(
                    konto="KRS↔ZOiS", punkt="Kapitał zakładowy",
                    status=StatusAudytu.BRAK,
                    uwagi="Brak w ZOiS konta z opisem 'kapitał zakładowy/podstawowy' "
                          "i brak Bilansu do porównania.",
                ))

        # ── Weryfikacja nazwy podmiotu (fuzzy) ──────────────────────────────
        nazwa_krs = dane_krs.nazwa.upper().strip()
        # TODO (opcjonalnie): porównanie z nazwą z nagłówka Bilansu/RZiS
        # Na razie tylko wyświetlamy nazwę z KRS jako INFO.


# =============================================================================
# INTEGRACJA Z API KRS (v3.2)
# =============================================================================

def pobierz_dane_krs(numer_krs: str, timeout_sec: int = 15) -> DaneKRS:
    """
    Pobiera dane spółki z oficjalnego API Ministerstwa Sprawiedliwości.
    Endpoint: https://api-krs.ms.gov.pl/api/krs/OdpisAktualny/{krs}?rejestr=P&format=json

    Args:
        numer_krs: 10-cyfrowy numer KRS (z lub bez wiodących zer)
        timeout_sec: timeout requesta (default 15s)

    Returns:
        DaneKRS: w przypadku sukcesu pełne dane, w przypadku błędu dane z
                 wypełnionym polem `blad` (resztę pól pustych).
    """
    import requests  # lazy import aby nie wymuszać zależności jeśli niepotrzebna

    # Normalizacja: tylko cyfry, padding do 10
    krs_clean = re.sub(r"[^0-9]", "", numer_krs or "")
    if not krs_clean:
        return DaneKRS(blad="Nie podano numeru KRS.")
    if len(krs_clean) > 10:
        return DaneKRS(blad=f"Numer KRS powinien mieć max 10 cyfr (podano {len(krs_clean)}).")
    krs_padded = krs_clean.zfill(10)

    url = f"https://api-krs.ms.gov.pl/api/krs/OdpisAktualny/{krs_padded}"
    params = {"rejestr": "P", "format": "json"}
    headers = {
        "Accept": "application/json",
        "User-Agent": "SymfoniaYearEndAuditor/3.2",
    }

    try:
        r = requests.get(url, params=params, headers=headers, timeout=timeout_sec)
    except requests.exceptions.Timeout:
        return DaneKRS(numer_krs=krs_padded, blad="Timeout API KRS (>15s).")
    except requests.exceptions.ConnectionError:
        return DaneKRS(numer_krs=krs_padded, blad="Brak połączenia z API KRS.")
    except Exception as e:
        return DaneKRS(numer_krs=krs_padded, blad=f"Błąd połączenia: {e}")

    if r.status_code == 404:
        return DaneKRS(numer_krs=krs_padded, blad=f"Nie znaleziono podmiotu o KRS {krs_padded}.")
    if r.status_code != 200:
        return DaneKRS(numer_krs=krs_padded, blad=f"API zwróciło status {r.status_code}.")

    try:
        data = r.json()
    except Exception as e:
        return DaneKRS(numer_krs=krs_padded, blad=f"Błąd parsowania JSON: {e}")

    # Nawigacja po strukturze (zgodnie z dokumentacją API KRS)
    try:
        dzial1 = data.get("odpis", {}).get("dane", {}).get("dzial1", {})
        dane_podm = dzial1.get("danePodmiotu", {})
        ident = dane_podm.get("identyfikatory", {})
        kapital_node = dzial1.get("kapital", {})

        # Kapitał zakładowy – format "5000,00 PLN" lub "5.000,00 PLN"
        kapital_str = kapital_node.get("wysokoscKapitaluZakladowego", "0")
        # Wyciąga tylko liczbę
        m = re.search(r"([\d.,]+)", str(kapital_str))
        if m:
            try:
                kapital = normalize_currency(m.group(1))
            except ValueError:
                kapital = Decimal("0")
        else:
            kapital = Decimal("0")

        from datetime import datetime
        return DaneKRS(
            numer_krs=krs_padded,
            nazwa=dane_podm.get("nazwa", "").strip(),
            forma_prawna=dane_podm.get("formaPrawna", "").strip(),
            nip=ident.get("nip", "").strip(),
            regon=(ident.get("regon", "") or "").strip()[:9],  # REGON 9-cyfrowy
            kapital_zakladowy=kapital,
            data_pobrania=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            blad=None,
        )
    except Exception as e:
        return DaneKRS(numer_krs=krs_padded, blad=f"Błąd parsowania odpowiedzi KRS: {e}")


# =============================================================================
# WYSYŁKA RAPORTU EMAIL (v3.4)
# =============================================================================

def wyslij_raport_email(
    *,
    nadawca: str,
    odbiorca: str,
    haslo: str,
    temat: str,
    tresc_tekstowa: str,
    serwer_smtp: str = "mail.abacus24.pl",
    port: int = 465,
) -> Tuple[bool, str]:
    """
    Wysyła raport audytu na podany adres e-mail przez SMTP SSL.

    Args:
        nadawca: adres skrzynki nadawczej (np. spraw_przyg@abacus24.pl)
        odbiorca: adres odbiorcy
        haslo: hasło do skrzynki nadawczej (zwykle ze st.secrets)
        temat: temat wiadomości
        tresc_tekstowa: treść jako plain text (raport + komentarze)
        serwer_smtp: hostname serwera SMTP (default mail.abacus24.pl)
        port: port SMTP (default 465 dla SSL)

    Returns:
        (sukces, komunikat): (True, "") przy sukcesie, (False, "...") przy błędzie
    """
    import smtplib
    import ssl
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.utils import formatdate, make_msgid

    if not nadawca or not odbiorca or not haslo:
        return False, "Brak nadawcy, odbiorcy lub hasła."

    try:
        msg = MIMEMultipart("alternative")
        msg["From"] = nadawca
        msg["To"] = odbiorca
        msg["Subject"] = temat
        msg["Date"] = formatdate(localtime=True)
        msg["Message-ID"] = make_msgid(domain=nadawca.split("@")[-1])
        msg.attach(MIMEText(tresc_tekstowa, "plain", "utf-8"))

        kontekst = ssl.create_default_context()
        with smtplib.SMTP_SSL(serwer_smtp, port, context=kontekst, timeout=30) as s:
            s.login(nadawca, haslo)
            s.sendmail(nadawca, [odbiorca], msg.as_string())

        return True, "Wiadomość wysłana pomyślnie."

    except smtplib.SMTPAuthenticationError:
        return False, "Błąd autoryzacji SMTP – sprawdź hasło w Streamlit Secrets."
    except smtplib.SMTPConnectError as e:
        return False, f"Nie udało się połączyć z {serwer_smtp}:{port} – {e}"
    except smtplib.SMTPException as e:
        return False, f"Błąd SMTP: {e}"
    except TimeoutError:
        return False, f"Timeout połączenia z {serwer_smtp}:{port}."
    except Exception as e:
        return False, f"Nieoczekiwany błąd wysyłki: {e}"




# =============================================================================
# KONIEC MODUŁU
# =============================================================================
# Ten plik jest modułem biblioteki – nie powinien być uruchamiany bezpośrednio.
# Aplikacja Streamlit (app.py) importuje z niego klasy i funkcje.
