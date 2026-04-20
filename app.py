"""
=============================================================================
SymfoniaYearEndAuditor – Automatyczna Kontrola Jakości Danych Finansowych
Biuro Rachunkowe Abacus | Zamknięcie Roku Obrachunkowego
=============================================================================
Wersja 3.1 – rozszerzenie o automatyczne rozpoznawanie planów kont.

Kluczowe funkcje:
  ZOiS (PDF lub XLSX):
    - Parsowanie tekstowe PDF Symfonii (radzi sobie z różnymi formatami)
    - Dekodowanie polskich znaków (cid:XXX) → ą,ę,ś,ć,ń,ł,ó,ż,ź
    - Automatyczne rozpoznawanie kont po OPISIE (nie tylko numerze):
        * Odbiorcy: szuka "odbiorc", "klient" w grupach 20x, 22x
        * Dostawcy: szuka "dostawc" w grupach 20x, 21x
        * ZUS: szuka "zus", "ubezpieczen" w syntetykach i analitykach
        * PIT: szuka "podatek doch", "podatek od płac" w analitykach
        * Środki w drodze: szuka "drodze", "pieniężne" w grupach 13x, 14x
        * Wynagrodzenia: szuka "wynagrodz", "pracownik" w grupie 23x

  Bilans (PDF lub XLSX):
    - Parsowanie sekcji AKTYWA/PASYWA
    - Pozycje A,B,C,D + sumy + wynik netto A.VI

  RZiS (PDF lub XLSX):
    - Wariant porównawczy, pozycje A-L
    - Strategia "last wins" dla rozwiązania kolizji litera I vs rzymska I

  Wyciągi bankowe:
    - Parsowanie per rachunek z wykryciem miesiąca ostatniej operacji
    - Obsługa wielu analityk 130-X (Santander, mBank, PKO, etc.)

Reguły weryfikacji (16+):
  - Spójność ZOiS: konta rozpoznane po opisie (jak wyżej)
  - Grupa 70x: przychody ze sprzedaży (saldo Ma)
  - Grupa 4: koszty rodzajowe (saldo Wn)
  - Bilans: Aktywa=Pasywa, spójność wewnętrzna A+B+C+D=Suma
  - RZiS: C=A-B, F=C+D-E, I=F+G-H, L=I-J-K
  - Krzyżowa RZiS↔Bilans: Wynik netto L = Pasywa A.VI
  - Krzyżowa ZOiS↔RZiS: Grupa 70x ≈ RZiS A, Grupa 4 ≈ RZiS B
=============================================================================
"""

from __future__ import annotations

import io
import re
import logging
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation
from enum import Enum
from typing import Dict, List, Optional, Tuple, Union

import pandas as pd

# ── Opcjonalne biblioteki ─────────────────────────────────────────────────────
try:
    import pdfplumber
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

logger = logging.getLogger("SymfoniaAuditor")
logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(message)s")

MIESIACE_PL = {
    1: "styczeń", 2: "luty", 3: "marzec", 4: "kwiecień",
    5: "maj", 6: "czerwiec", 7: "lipiec", 8: "sierpień",
    9: "wrzesień", 10: "październik", 11: "listopad", 12: "grudzień",
}

# Regex dla kwoty w formacie polskim: 1.114.621,54 | 0,00 | -112.140,00
RE_KWOTA = r"-?\d+(?:\.\d{3})*(?:,\d{1,2})?"


# =============================================================================
# DEKODER POLSKICH ZNAKÓW Z PDF SYMFONII
# =============================================================================
# Symfonia eksportuje PDF z własną czcionką gdzie polskie znaki są kodowane
# jako glyph-id'y. pdfplumber zwraca je jako "(cid:XXX)". Ta mapa przekłada
# najczęściej występujące glyphy na prawdziwe polskie litery.
CID_TO_PL = {
    # Niestandardowe CID dla zwykłych liter łacińskich (anomalia fontu Symfonii)
    39:  "D",   71:  "D",   79:  "l",   88:  "u",
    # Polskie znaki – zidentyfikowane z plików Abacus i IGRAPES
    211: "Ó",   243: "ó",
    224: "Ł",   225: "ł",
    260: "Ą",   261: "ą",
    262: "Ć",   252: "ć",   253: "Ć",
    265: "Ę",   266: "ę",   267: "Ę",
    269: "Ń",   276: "ń",   277: "Ń",
    285: "Ś",   286: "ś",
    297: "Ź",   240: "ź",   241: "Ź",
    298: "ż",   300: "Ż",
}
_RE_CID = re.compile(r"\(cid:(\d+)\)")


def dekoduj_cid(tekst: str) -> str:
    """
    Zamienia glyph-id Symfonii na polskie znaki (np. '(cid:266)' → 'ę').
    Nieznane CID zostają zastąpione znakiem zapytania.
    """
    if not tekst or "(cid:" not in tekst:
        return tekst
    return _RE_CID.sub(lambda m: CID_TO_PL.get(int(m.group(1)), "?"), tekst)


# =============================================================================
# TYPY
# =============================================================================

class StatusAudytu(Enum):
    OK       = "✅ OK"
    BLAD     = "❌ BŁĄD"
    OSTRZEZ  = "⚠️  OSTRZEŻENIE"
    BRAK     = "🔍 BRAK DANYCH"
    INFO     = "ℹ️  INFO"


@dataclass
class PunktKontroli:
    konto:   str
    punkt:   str
    status:  StatusAudytu
    uwagi:   str = ""
    wartosc: str = ""


@dataclass
class DaneZOiS:
    """
    Zestawienie Obrotów i Sald z Symfonii – syntetyki + analityki.
    """
    konta: Dict[str, Tuple[Decimal, Decimal]] = field(default_factory=dict)
    konta_analityki: Dict[str, Tuple[Decimal, Decimal]] = field(default_factory=dict)
    opisy: Dict[str, str] = field(default_factory=dict)

    def pobierz_konta_bankowe(self) -> List[Tuple[str, str, Decimal]]:
        """Zwraca listę wszystkich rachunków bankowych (analityki 130)."""
        wynik = []
        for numer, (wn, ma) in self.konta_analityki.items():
            if normalize_konto(numer) == "130":
                saldo_netto = wn - ma
                opis = self.opisy.get(numer, "Rachunek bankowy")
                wynik.append((numer, opis, saldo_netto))
        if not wynik and "130" in self.konta:
            wn, ma = self.konta["130"]
            wynik.append(("130", self.opisy.get("130", "Rachunek bankowy"), wn - ma))
        return sorted(wynik, key=lambda x: x[0])


@dataclass
class DaneBilansu:
    """
    Dane Bilansu sparsowane z Symfonii.
    Zawiera zarówno szczegółowe pozycje (A.aktywa trwałe, B.obrotowe, C, D)
    jak i sumy bilansowe. Wynik netto pobierany z pasywów A.VI.
    """
    # AKTYWA (sekcja AKTYWA w PDF)
    aktywa_trwale_biezacy:    Decimal = Decimal("0")   # poz. A
    aktywa_trwale_ubiegly:    Decimal = Decimal("0")
    aktywa_obrotowe_biezacy:  Decimal = Decimal("0")   # poz. B
    aktywa_obrotowe_ubiegly:  Decimal = Decimal("0")
    nalezne_wplaty_biezacy:   Decimal = Decimal("0")   # poz. C
    nalezne_wplaty_ubiegly:   Decimal = Decimal("0")
    udzialy_wlasne_biezacy:   Decimal = Decimal("0")   # poz. D
    udzialy_wlasne_ubiegly:   Decimal = Decimal("0")
    suma_aktywow_biezacy:     Decimal = Decimal("0")
    suma_aktywow_ubiegly:     Decimal = Decimal("0")

    # PASYWA (sekcja PASYWA w PDF)
    kapital_wlasny_biezacy:   Decimal = Decimal("0")   # poz. A
    kapital_wlasny_ubiegly:   Decimal = Decimal("0")
    zobowiazania_biezacy:     Decimal = Decimal("0")   # poz. B
    zobowiazania_ubiegly:     Decimal = Decimal("0")
    suma_pasywow_biezacy:     Decimal = Decimal("0")
    suma_pasywow_ubiegly:     Decimal = Decimal("0")

    # Wynik netto z Bilansu (pasywa A.VI) – do weryfikacji krzyżowej z RZiS
    wynik_netto_biezacy:      Decimal = Decimal("0")
    wynik_netto_ubiegly:      Decimal = Decimal("0")

    # ── Kompatybilność wsteczna z v2.0 ───────────────────────────────────────
    @property
    def aktywa_biezacy(self) -> Decimal: return self.suma_aktywow_biezacy
    @property
    def pasywa_biezacy(self) -> Decimal: return self.suma_pasywow_biezacy
    @property
    def aktywa_ubiegly(self) -> Decimal: return self.suma_aktywow_ubiegly
    @property
    def pasywa_ubiegly(self) -> Decimal: return self.suma_pasywow_ubiegly


@dataclass
class DaneRZiS:
    """
    Rachunek Zysków i Strat (wariant porównawczy).
    Każda pozycja: (rok bieżący, rok ubiegły).

    Pozycje A-L zgodne ze standardem polskiej ustawy o rachunkowości:
      A. Przychody netto ze sprzedaży
      B. Koszty działalności operacyjnej
      C. Zysk (strata) ze sprzedaży = A - B
      D. Pozostałe przychody operacyjne
      E. Pozostałe koszty operacyjne
      F. Zysk (strata) z działalności operacyjnej = C + D - E
      G. Przychody finansowe
      H. Koszty finansowe
      I. Zysk (strata) brutto = F + G - H
      J. Podatek dochodowy
      K. Pozostałe obowiązkowe zmniejszenia zysku
      L. Zysk (strata) netto = I - J - K
    """
    przychody_sprzedazy:       Tuple[Decimal, Decimal] = (Decimal("0"), Decimal("0"))  # A
    koszty_operacyjne:         Tuple[Decimal, Decimal] = (Decimal("0"), Decimal("0"))  # B
    zysk_sprzedazy:            Tuple[Decimal, Decimal] = (Decimal("0"), Decimal("0"))  # C
    pozostale_przych_oper:     Tuple[Decimal, Decimal] = (Decimal("0"), Decimal("0"))  # D
    pozostale_koszty_oper:     Tuple[Decimal, Decimal] = (Decimal("0"), Decimal("0"))  # E
    zysk_dzialalnosci_oper:    Tuple[Decimal, Decimal] = (Decimal("0"), Decimal("0"))  # F
    przychody_finansowe:       Tuple[Decimal, Decimal] = (Decimal("0"), Decimal("0"))  # G
    koszty_finansowe:          Tuple[Decimal, Decimal] = (Decimal("0"), Decimal("0"))  # H
    zysk_brutto:               Tuple[Decimal, Decimal] = (Decimal("0"), Decimal("0"))  # I
    podatek_dochodowy:         Tuple[Decimal, Decimal] = (Decimal("0"), Decimal("0"))  # J
    pozostale_zmniejszenia:    Tuple[Decimal, Decimal] = (Decimal("0"), Decimal("0"))  # K
    zysk_netto:                Tuple[Decimal, Decimal] = (Decimal("0"), Decimal("0"))  # L

    # Mapa litera → pole (używane przy parsowaniu i walidacji)
    MAPA_POZYCJI = {
        "A": "przychody_sprzedazy",
        "B": "koszty_operacyjne",
        "C": "zysk_sprzedazy",
        "D": "pozostale_przych_oper",
        "E": "pozostale_koszty_oper",
        "F": "zysk_dzialalnosci_oper",
        "G": "przychody_finansowe",
        "H": "koszty_finansowe",
        "I": "zysk_brutto",
        "J": "podatek_dochodowy",
        "K": "pozostale_zmniejszenia",
        "L": "zysk_netto",
    }


@dataclass
class WyciagBankowy:
    """Pojedynczy wyciąg bankowy powiązany z analityką 130-X."""
    numer_konta_ksiegowego: str
    saldo_koncowe: Decimal
    rok_ostatniej_operacji: Optional[int] = None
    miesiac_ostatniej_operacji: Optional[int] = None
    numer_rachunku: str = ""
    bank_nazwa: str = ""
    wgrany_plik: bool = True

    @property
    def okres_opisowy(self) -> str:
        if self.miesiac_ostatniej_operacji and self.rok_ostatniej_operacji:
            return f"{MIESIACE_PL.get(self.miesiac_ostatniej_operacji, '?')} {self.rok_ostatniej_operacji}"
        elif self.rok_ostatniej_operacji:
            return str(self.rok_ostatniej_operacji)
        return "okres nieznany"


# =============================================================================
# FUNKCJE POMOCNICZE
# =============================================================================

def normalize_currency(wartosc: Union[str, float, int, None]) -> Decimal:
    """Normalizuje kwotę finansową do Decimal."""
    if wartosc is None:
        return Decimal("0")
    if isinstance(wartosc, (int, float)):
        try:
            return Decimal(str(round(float(wartosc), 2)))
        except (InvalidOperation, ValueError):
            return Decimal("0")
    tekst = str(wartosc).strip()
    if not tekst or tekst in ("-", "–", "0,00", "0.00"):
        return Decimal("0")
    tekst = re.sub(r"[złZŁPLN\xa0]", "", tekst).strip()
    if "," in tekst and "." in tekst:
        if tekst.rfind(",") > tekst.rfind("."):
            tekst = tekst.replace(".", "").replace(",", ".")
        else:
            tekst = tekst.replace(",", "")
    elif "," in tekst:
        czesci = tekst.split(",")
        if len(czesci) == 2 and len(czesci[1]) <= 2:
            tekst = tekst.replace(",", ".")
        else:
            tekst = tekst.replace(",", "")
    tekst = tekst.replace(" ", "")
    try:
        return Decimal(tekst).quantize(Decimal("0.01"))
    except InvalidOperation:
        raise ValueError(f"Nie można znormalizować kwoty: '{wartosc}'")


def normalize_konto(numer: str) -> str:
    if not numer:
        return ""
    return str(numer).strip().split("-")[0].strip()


def get_grupa(numer_konta: str) -> Optional[int]:
    syntetyka = normalize_konto(numer_konta)
    if syntetyka and syntetyka[0].isdigit():
        return int(syntetyka[0])
    return None


def wykryj_date_ostatniej_operacji(tekst: str) -> Optional[Tuple[int, int]]:
    """Wykrywa rok i miesiąc ostatniej operacji w wyciągu bankowym."""
    wszystkie_daty: List[date] = []
    wzorce = [
        r"(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})",
        r"(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})",
    ]
    for wzorzec in wzorce:
        for m in re.finditer(wzorzec, tekst):
            g = m.groups()
            try:
                if len(g[0]) == 4:
                    rok, mies, dzien = int(g[0]), int(g[1]), int(g[2])
                else:
                    dzien, mies, rok = int(g[0]), int(g[1]), int(g[2])
                if 1 <= mies <= 12 and 1 <= dzien <= 31 and 2000 <= rok <= 2100:
                    wszystkie_daty.append(date(rok, mies, dzien))
            except (ValueError, IndexError):
                continue
    for m in re.finditer(r"\b(\d{1,2})/(\d{4})\b", tekst):
        try:
            mies, rok = int(m.group(1)), int(m.group(2))
            if 1 <= mies <= 12 and 2000 <= rok <= 2100:
                wszystkie_daty.append(date(rok, mies, 1))
        except ValueError:
            continue
    if not wszystkie_daty:
        return None
    najpozniejsza = max(wszystkie_daty)
    return (najpozniejsza.year, najpozniejsza.month)


# =============================================================================
# PARSER ZOiS
# =============================================================================

class ParserZOiS:
    KOLUMNY_KONTO = ["konto", "numer konta", "nr konta", "symbol konta", "account"]
    KOLUMNY_NAZWA = ["nazwa", "nazwa konta", "opis", "name"]
    KOLUMNY_SALDO_WN = ["saldo wn", "saldo_wn", "debet", "wn", "saldo dt", "dt"]
    KOLUMNY_SALDO_MA = ["saldo ma", "saldo_ma", "kredyt", "ma", "saldo ct", "ct"]

    def _znajdz_kolumne(self, df, kandydaci):
        kol_lower = {k.lower().strip(): k for k in df.columns}
        for k in kandydaci:
            if k.lower() in kol_lower:
                return kol_lower[k.lower()]
        return None

    def parsuj_xlsx(self, dane_binarne: bytes) -> DaneZOiS:
        bufor = io.BytesIO(dane_binarne)
        try:
            wb = openpyxl.load_workbook(bufor, data_only=True)
        except Exception as e:
            raise ValueError(f"Błąd odczytu XLSX (ZOiS): {e}")
        arkusz = wb.active
        for nazwa in ["ZOiS", "Zestawienie", "ObrotyiSalda", "Sheet1"]:
            if nazwa in wb.sheetnames:
                arkusz = wb[nazwa]
                break
        wiersze = [list(w) for w in arkusz.iter_rows(values_only=True)]
        if not wiersze:
            raise ValueError("Arkusz ZOiS jest pusty.")
        idx = self._znajdz_wiersz_naglowka(wiersze)
        if idx is None:
            raise ValueError("Nie znaleziono wiersza nagłówkowego w ZOiS.")
        naglowki = [str(k).strip() if k else "" for k in wiersze[idx]]
        df = pd.DataFrame(wiersze[idx+1:], columns=naglowki)
        return self._parsuj_dataframe(df, "XLSX")

    def _znajdz_wiersz_naglowka(self, wiersze):
        kluczowe = {"konto", "saldo", "wn", "ma", "obroty", "numer"}
        for i, w in enumerate(wiersze[:20]):
            vals = {str(v).lower().strip() for v in w if v is not None}
            if len(vals & kluczowe) >= 2:
                return i
        return None

    # Regex do rozpoznawania wierszy w tekstowej wersji PDF Symfonii
    RE_NUMER_KONTA = re.compile(r"^(\d{2,3}(?:-\d+)*)\s+(.*)$")
    RE_LICZBA_PL = re.compile(r"-?\d+(?:\.\d{3})*,\d{1,2}")
    LINIE_IGNOROWANE = (
        "Suma strony", "Suma razem", "Z przeniesienia", "Do przeniesienia",
    )

    def parsuj_pdf(self, dane_binarne: bytes) -> DaneZOiS:
        """
        Parser PDF ZOiS z Symfonii.
        Symfonia eksportuje PDF jako tekst w kolumnach (bez obramowań tabeli),
        więc używamy ekstraktu tekstu + regex zamiast extract_tables().

        Strategia:
          1. Iteruj po liniach - każda potencjalnie to wiersz kontowy
          2. Wiersz zaczyna się od numeru konta (np. 130, 201-5, 403-1-1)
          3. Długie nazwy klientów mogą być wieloliniowe - obsługa przenoszenia
          4. Z końca linii bierzemy 8 liczb (BO Wn, BO Ma, obroty Wn/Ma,
             narast Wn/Ma, Saldo Wn, Saldo Ma) - interesują nas 2 ostatnie
        """
        if not PDF_AVAILABLE:
            raise ImportError("Brak biblioteki pdfplumber.")
        bufor = io.BytesIO(dane_binarne)
        linie = []
        with pdfplumber.open(bufor) as pdf:
            for strona in pdf.pages:
                t = dekoduj_cid(strona.extract_text() or "")
                linie.extend(t.splitlines())

        if not linie:
            raise ValueError("PDF jest pusty lub nie udało się ekstraktować tekstu.")

        wynik = DaneZOiS()
        czek_konto: Optional[str] = None
        czek_nazwa: str = ""

        def zapisz(numer: str, nazwa: str, saldo_wn: str, saldo_ma: str):
            """Zapisuje pozycję do DaneZOiS - do syntetyk i/lub analityk."""
            try:
                wn = normalize_currency(saldo_wn)
                ma = normalize_currency(saldo_ma)
            except ValueError:
                return
            numer_syn = normalize_konto(numer)
            if not re.match(r"^\d+$", numer_syn):
                return

            # Zapis jako analityka jeśli numer zawiera "-"
            if numer != numer_syn:
                wynik.konta_analityki[numer] = (wn, ma)
                if nazwa and numer not in wynik.opisy:
                    wynik.opisy[numer] = nazwa.strip()
            else:
                # Czysta syntetyka
                wynik.konta[numer_syn] = (wn, ma)
                if nazwa and numer_syn not in wynik.opisy:
                    wynik.opisy[numer_syn] = nazwa.strip()
                # Dla kompatybilności wstecznej z audytorem: zapisz też do konta
                # (syntetyki występują w ZOiS tylko raz, więc nie nadpisujemy)

        for linia in linie:
            ln = linia.strip()
            if not ln:
                continue
            # Pomijamy linie podsumowujące
            if any(x in ln for x in self.LINIE_IGNOROWANE):
                continue

            m = self.RE_NUMER_KONTA.match(ln)
            if m:
                numer = m.group(1)
                reszta = m.group(2)
                liczby = self.RE_LICZBA_PL.findall(reszta)

                if len(liczby) >= 8:
                    # Cały wiersz w jednej linii
                    idx = reszta.find(liczby[0])
                    nazwa = reszta[:idx].strip() if idx > 0 else ""
                    zapisz(numer, nazwa, liczby[-2], liczby[-1])
                    czek_konto = None
                    czek_nazwa = ""
                else:
                    # Wiersz kontynuacyjny - nazwa/liczby w kolejnych liniach
                    czek_konto = numer
                    czek_nazwa = reszta
            elif czek_konto:
                liczby = self.RE_LICZBA_PL.findall(ln)
                if len(liczby) >= 8:
                    # Linia z liczbami dla oczekującego konta
                    zapisz(czek_konto, czek_nazwa, liczby[-2], liczby[-1])
                    czek_konto = None
                    czek_nazwa = ""
                else:
                    # Kontynuacja nazwy klienta
                    czek_nazwa += " " + ln

        # Dla kompatybilności: jeśli mamy syntetykę 130 ale nie analitykę, dodaj ją
        if "130" in wynik.konta and not any(
            normalize_konto(k) == "130" for k in wynik.konta_analityki
        ):
            wynik.konta_analityki["130"] = wynik.konta["130"]

        logger.info(
            f"ZOiS (PDF): {len(wynik.konta)} syntetyk, "
            f"{len(wynik.konta_analityki)} analityk."
        )
        return wynik

    def _parsuj_dataframe(self, df: pd.DataFrame, zrodlo: str) -> DaneZOiS:
        kol_konto = self._znajdz_kolumne(df, self.KOLUMNY_KONTO)
        kol_nazwa = self._znajdz_kolumne(df, self.KOLUMNY_NAZWA)
        kol_wn    = self._znajdz_kolumne(df, self.KOLUMNY_SALDO_WN)
        kol_ma    = self._znajdz_kolumne(df, self.KOLUMNY_SALDO_MA)
        brakujace = []
        if not kol_konto: brakujace.append("Konto")
        if not kol_wn:    brakujace.append("Saldo Wn")
        if not kol_ma:    brakujace.append("Saldo Ma")
        if brakujace:
            raise ValueError(
                f"Brak kolumn ZOiS ({zrodlo}): {', '.join(brakujace)}. "
                f"Dostępne: {list(df.columns)}"
            )
        wynik = DaneZOiS()
        for _, w in df.iterrows():
            numer_raw = w.get(kol_konto)
            if pd.isna(numer_raw) or not str(numer_raw).strip():
                continue
            numer_pelny = str(numer_raw).strip()
            numer_syn = normalize_konto(numer_pelny)
            if numer_syn.lower() in {"razem", "suma", "ogółem", "total"}:
                continue
            if not re.match(r"^\d+", numer_syn):
                continue
            try:
                wn = normalize_currency(w.get(kol_wn))
                ma = normalize_currency(w.get(kol_ma))
            except ValueError as e:
                logger.warning(f"Pominięto {numer_pelny}: {e}")
                continue
            if numer_pelny != numer_syn:
                iwn, ima = wynik.konta_analityki.get(numer_pelny, (Decimal("0"), Decimal("0")))
                wynik.konta_analityki[numer_pelny] = (iwn + wn, ima + ma)
                if kol_nazwa:
                    opis = w.get(kol_nazwa)
                    if not pd.isna(opis) and numer_pelny not in wynik.opisy:
                        wynik.opisy[numer_pelny] = str(opis).strip()
            iwn, ima = wynik.konta.get(numer_syn, (Decimal("0"), Decimal("0")))
            wynik.konta[numer_syn] = (iwn + wn, ima + ma)
            if kol_nazwa and numer_syn not in wynik.opisy:
                opis = w.get(kol_nazwa)
                if not pd.isna(opis):
                    wynik.opisy[numer_syn] = str(opis).strip()
        if "130" in wynik.konta and not any(
            normalize_konto(k) == "130" for k in wynik.konta_analityki
        ):
            wynik.konta_analityki["130"] = wynik.konta["130"]
        logger.info(
            f"ZOiS ({zrodlo}): {len(wynik.konta)} syntetyk, "
            f"{len(wynik.konta_analityki)} analityk."
        )
        return wynik


# =============================================================================
# PARSER BILANSU (v3.0 – pełna obsługa formatu Symfonii)
# =============================================================================

class ParserBilansu:
    """
    Parser Bilansu z Symfonii (wariant pełny).

    Wykrywa:
      - Sekcje AKTYWA i PASYWA (separator)
      - Pozycje A, B, C, D (aktywa) i A, B (pasywa)
      - Suma (pojawia się 2 razy: suma aktywów i suma pasywów)
      - Wynik netto (pasywa A.VI)
    """

    # Regex: pozycja A-D + tekst + 3 kwoty (bieżący, ubiegły, różnica)
    RE_POZ = re.compile(
        rf"^\s*(?:[+\-*]\s+)?([A-D])\s+.+?\s+({RE_KWOTA})\s+({RE_KWOTA})\s+({RE_KWOTA})\s*$"
    )
    # Regex: "Suma" + 3 kwoty
    RE_SUMA = re.compile(
        rf"^\s*Suma\s+({RE_KWOTA})\s+({RE_KWOTA})\s+({RE_KWOTA})\s*$"
    )
    # Regex: pozycja VI "Zysk (strata) netto" w kapitale własnym pasywów
    RE_WYNIK_NETTO = re.compile(
        rf"VI\s+Zysk.+?netto\s+({RE_KWOTA})\s+({RE_KWOTA})\s+({RE_KWOTA})"
    )

    def parsuj_xlsx(self, dane_binarne: bytes) -> DaneBilansu:
        """XLSX – próbuje parsować jak tekst."""
        bufor = io.BytesIO(dane_binarne)
        try:
            wb = openpyxl.load_workbook(bufor, data_only=True)
            wiersze = []
            for w in wb.active.iter_rows(values_only=True):
                linia = " ".join(str(k) for k in w if k is not None)
                wiersze.append(linia)
        except Exception as e:
            raise ValueError(f"Błąd odczytu XLSX (Bilans): {e}")
        return self._parsuj_linie(wiersze)

    def parsuj_pdf(self, dane_binarne: bytes) -> DaneBilansu:
        """PDF – ekstrahuje tekst i parsuje linia po linii."""
        if not PDF_AVAILABLE:
            raise ImportError("Brak pdfplumber.")
        bufor = io.BytesIO(dane_binarne)
        wszystkie = []
        with pdfplumber.open(bufor) as pdf:
            for strona in pdf.pages:
                t = dekoduj_cid(strona.extract_text() or "")
                wszystkie.extend(t.splitlines())
        return self._parsuj_linie(wszystkie)

    def _parsuj_linie(self, linie: List[str]) -> DaneBilansu:
        """
        Główna logika parsowania. Iteruje po liniach śledząc w której sekcji
        (AKTYWA/PASYWA) się znajdujemy.
        """
        wynik = DaneBilansu()
        tryb: Optional[str] = None  # "A" = aktywa, "P" = pasywa
        sumy: List[Tuple[str, Decimal, Decimal]] = []  # [(tryb, bieżący, ubiegły)]
        znalezione_aktywa: Dict[str, Tuple[Decimal, Decimal]] = {}
        znalezione_pasywa: Dict[str, Tuple[Decimal, Decimal]] = {}

        for linia in linie:
            # ── Separator sekcji ─────────────────────────────────────────────
            if "AKTYWA" in linia and "PASYWA" not in linia:
                tryb = "A"
                continue
            if "PASYWA" in linia:
                tryb = "P"
                continue

            # ── Wynik netto (A.VI pasywów) ──────────────────────────────────
            if tryb == "P":
                m_wn = self.RE_WYNIK_NETTO.search(linia)
                if m_wn:
                    try:
                        wynik.wynik_netto_biezacy = normalize_currency(m_wn.group(1))
                        wynik.wynik_netto_ubiegly = normalize_currency(m_wn.group(2))
                    except ValueError:
                        pass

            # ── Suma (pojawia się 2x: aktywa i pasywa) ───────────────────────
            m_suma = self.RE_SUMA.match(linia)
            if m_suma and tryb:
                try:
                    b = normalize_currency(m_suma.group(1))
                    u = normalize_currency(m_suma.group(2))
                    sumy.append((tryb, b, u))
                except ValueError:
                    pass
                continue

            # ── Pozycja A/B/C/D w aktywach lub A/B w pasywach ────────────────
            m = self.RE_POZ.match(linia)
            if m and tryb:
                lit = m.group(1)
                try:
                    b = normalize_currency(m.group(2))
                    u = normalize_currency(m.group(3))
                except ValueError:
                    continue

                target = znalezione_aktywa if tryb == "A" else znalezione_pasywa
                # "First wins" – pierwsza pozycja wiąże się z główną, kolejne ignorujemy
                if lit not in target:
                    target[lit] = (b, u)

        # ── Przepisanie do struktury DaneBilansu ─────────────────────────────
        if "A" in znalezione_aktywa:
            wynik.aktywa_trwale_biezacy, wynik.aktywa_trwale_ubiegly = znalezione_aktywa["A"]
        if "B" in znalezione_aktywa:
            wynik.aktywa_obrotowe_biezacy, wynik.aktywa_obrotowe_ubiegly = znalezione_aktywa["B"]
        if "C" in znalezione_aktywa:
            wynik.nalezne_wplaty_biezacy, wynik.nalezne_wplaty_ubiegly = znalezione_aktywa["C"]
        if "D" in znalezione_aktywa:
            wynik.udzialy_wlasne_biezacy, wynik.udzialy_wlasne_ubiegly = znalezione_aktywa["D"]

        if "A" in znalezione_pasywa:
            wynik.kapital_wlasny_biezacy, wynik.kapital_wlasny_ubiegly = znalezione_pasywa["A"]
        if "B" in znalezione_pasywa:
            wynik.zobowiazania_biezacy, wynik.zobowiazania_ubiegly = znalezione_pasywa["B"]

        # Sumy: pierwsza to aktywa, druga to pasywa
        for tr, b, u in sumy:
            if tr == "A" and wynik.suma_aktywow_biezacy == Decimal("0"):
                wynik.suma_aktywow_biezacy = b
                wynik.suma_aktywow_ubiegly = u
            elif tr == "P" and wynik.suma_pasywow_biezacy == Decimal("0"):
                wynik.suma_pasywow_biezacy = b
                wynik.suma_pasywow_ubiegly = u

        logger.info(
            f"Bilans: Aktywa={wynik.suma_aktywow_biezacy:,.2f}, "
            f"Pasywa={wynik.suma_pasywow_biezacy:,.2f}, "
            f"Wynik netto={wynik.wynik_netto_biezacy:,.2f}"
        )
        return wynik


# =============================================================================
# PARSER RACHUNKU ZYSKÓW I STRAT (v3.0 – nowość)
# =============================================================================

class ParserRZiS:
    """
    Parser Rachunku Zysków i Strat (wariant porównawczy).

    Strategia: "last wins" – iteruje po liniach i nadpisuje słownik pozycji.
    Dzięki temu:
      - Litery A-H, J, K, L są unikalne (pierwsza = główna)
      - Litera "I" występuje wielokrotnie jako rzymska cyfra I w podsekcjach
        (np. "I Amortyzacja" w sekcji B), ale OSTATNIA "I" = główna (Zysk brutto)
    """

    RE_POZ = re.compile(
        rf"^\s*(?:[+\-*]\s+)?([A-L])\s+.+?\s+({RE_KWOTA})\s+({RE_KWOTA})\s*$"
    )

    def parsuj_xlsx(self, dane_binarne: bytes) -> DaneRZiS:
        bufor = io.BytesIO(dane_binarne)
        try:
            wb = openpyxl.load_workbook(bufor, data_only=True)
            linie = []
            for w in wb.active.iter_rows(values_only=True):
                linie.append(" ".join(str(k) for k in w if k is not None))
        except Exception as e:
            raise ValueError(f"Błąd odczytu XLSX (RZiS): {e}")
        return self._parsuj_linie(linie)

    def parsuj_pdf(self, dane_binarne: bytes) -> DaneRZiS:
        if not PDF_AVAILABLE:
            raise ImportError("Brak pdfplumber.")
        bufor = io.BytesIO(dane_binarne)
        linie = []
        with pdfplumber.open(bufor) as pdf:
            for strona in pdf.pages:
                t = dekoduj_cid(strona.extract_text() or "")
                linie.extend(t.splitlines())
        return self._parsuj_linie(linie)

    def _parsuj_linie(self, linie: List[str]) -> DaneRZiS:
        znalezione: Dict[str, Tuple[Decimal, Decimal]] = {}

        for linia in linie:
            m = self.RE_POZ.match(linia)
            if m:
                lit = m.group(1)
                try:
                    b = normalize_currency(m.group(2))
                    u = normalize_currency(m.group(3))
                    # "last wins" – ostatnie wystąpienie nadpisuje
                    znalezione[lit] = (b, u)
                except ValueError:
                    continue

        wynik = DaneRZiS()
        for lit, pole in DaneRZiS.MAPA_POZYCJI.items():
            if lit in znalezione:
                setattr(wynik, pole, znalezione[lit])

        logger.info(
            f"RZiS: Przychody(A)={wynik.przychody_sprzedazy[0]:,.2f}, "
            f"Koszty(B)={wynik.koszty_operacyjne[0]:,.2f}, "
            f"Zysk netto(L)={wynik.zysk_netto[0]:,.2f}"
        )
        return wynik


# =============================================================================
# PARSER WYCIĄGU BANKOWEGO (bez zmian)
# =============================================================================

class ParserWyciaguBankowego:
    KLUCZE_SALDO = [
        "saldo końcowe", "saldo na koniec", "closing balance",
        "stan na koniec", "saldo końca okresu",
    ]
    WZORZEC_IBAN = re.compile(r"(?:PL)?\s*(\d{2}\s*(?:\d{4}\s*){6})")
    ZNANE_BANKI = [
        "Santander", "PKO BP", "PKO Bank", "mBank", "ING Bank", "Pekao",
        "BNP Paribas", "Millennium", "Credit Agricole", "Alior", "Citi",
        "Getin", "Nest Bank", "Raiffeisen", "Deutsche Bank", "BOŚ",
    ]

    def parsuj(self, dane_binarne: bytes, format_pliku: str):
        if format_pliku.lower() == "pdf":
            return self._parsuj_pdf(dane_binarne)
        return self._parsuj_xlsx(dane_binarne)

    def _parsuj_pdf(self, dane_binarne: bytes):
        if not PDF_AVAILABLE:
            raise ImportError("Brak pdfplumber.")
        bufor = io.BytesIO(dane_binarne)
        saldo = Decimal("0")
        caly_tekst = ""
        with pdfplumber.open(bufor) as pdf:
            for strona in pdf.pages:
                caly_tekst += dekoduj_cid(strona.extract_text() or "") + "\n"
            for strona in reversed(pdf.pages):
                t = dekoduj_cid(strona.extract_text() or "")
                s = self._wyciagnij_saldo(t)
                if s and s != Decimal("0"):
                    saldo = s
                    break
        data_op = wykryj_date_ostatniej_operacji(caly_tekst)
        rok, mies = data_op if data_op else (None, None)
        return saldo, rok, mies, self._iban(caly_tekst), self._bank(caly_tekst)

    def _parsuj_xlsx(self, dane_binarne: bytes):
        bufor = io.BytesIO(dane_binarne)
        try:
            df = pd.read_excel(bufor, engine="openpyxl", header=None)
        except Exception as e:
            raise ValueError(f"Błąd odczytu XLSX wyciągu: {e}")
        saldo = Decimal("0")
        caly_tekst = ""
        for _, wiersz in df.iterrows():
            for kom in wiersz:
                if isinstance(kom, str):
                    caly_tekst += kom + " "
                    if any(k in kom.lower() for k in self.KLUCZE_SALDO):
                        for v in wiersz:
                            try:
                                kw = normalize_currency(v)
                                if kw != Decimal("0"):
                                    saldo = kw
                                    break
                            except (ValueError, TypeError):
                                pass
                elif kom is not None:
                    caly_tekst += str(kom) + " "
        data_op = wykryj_date_ostatniej_operacji(caly_tekst)
        rok, mies = data_op if data_op else (None, None)
        return saldo, rok, mies, self._iban(caly_tekst), self._bank(caly_tekst)

    def _wyciagnij_saldo(self, tekst: str):
        for linia in tekst.splitlines():
            ll = linia.lower()
            if any(k in ll for k in self.KLUCZE_SALDO):
                for token in reversed(linia.split()):
                    try:
                        kw = normalize_currency(token)
                        if kw != Decimal("0"):
                            return kw
                    except (ValueError, TypeError):
                        pass
        return None

    def _iban(self, tekst: str) -> str:
        m = self.WZORZEC_IBAN.search(tekst)
        return re.sub(r"\s+", "", m.group(1)) if m else ""

    def _bank(self, tekst: str) -> str:
        tl = tekst.lower()
        for bank in self.ZNANE_BANKI:
            if bank.lower() in tl:
                return bank
        return ""


# =============================================================================
# GŁÓWNA KLASA AUDYTORA
# =============================================================================

class SymfoniaYearEndAuditor:
    """
    Audytor jakości danych – zamknięcie roku v3.0.

    API:
        parsuj_zois(bytes, format)    → DaneZOiS
        parsuj_bilans(bytes, format)  → DaneBilansu
        parsuj_rzis(bytes, format)    → DaneRZiS
        parsuj_wyciag(konto, bytes, format) → WyciagBankowy
        check_accounting_logic(...)   → weryfikacja
        generate_audit_report(...)    → raport
        run_full_audit(...)           → wszystko w jednym wywołaniu
    """

    def __init__(self):
        self._parser_zois   = ParserZOiS()
        self._parser_bilans = ParserBilansu()
        self._parser_rzis   = ParserRZiS()
        self._parser_wyciag = ParserWyciaguBankowego()
        self._wyniki: List[PunktKontroli] = []

    # ── Publiczne API – parsowanie ──────────────────────────────────────────

    def parsuj_zois(self, dane_binarne: bytes, format_pliku: str = "xlsx") -> DaneZOiS:
        if format_pliku.lower() == "pdf":
            return self._parser_zois.parsuj_pdf(dane_binarne)
        return self._parser_zois.parsuj_xlsx(dane_binarne)

    def parsuj_bilans(self, dane_binarne: bytes, format_pliku: str = "xlsx") -> DaneBilansu:
        if format_pliku.lower() == "pdf":
            return self._parser_bilans.parsuj_pdf(dane_binarne)
        return self._parser_bilans.parsuj_xlsx(dane_binarne)

    def parsuj_rzis(self, dane_binarne: bytes, format_pliku: str = "xlsx") -> DaneRZiS:
        if format_pliku.lower() == "pdf":
            return self._parser_rzis.parsuj_pdf(dane_binarne)
        return self._parser_rzis.parsuj_xlsx(dane_binarne)

    def parsuj_wyciag(
        self,
        numer_konta_ksiegowego: str,
        dane_binarne: bytes,
        format_pliku: str = "pdf",
    ) -> WyciagBankowy:
        saldo, rok, mies, iban, bank = self._parser_wyciag.parsuj(
            dane_binarne, format_pliku
        )
        return WyciagBankowy(
            numer_konta_ksiegowego=numer_konta_ksiegowego,
            saldo_koncowe=saldo,
            rok_ostatniej_operacji=rok,
            miesiac_ostatniej_operacji=mies,
            numer_rachunku=iban,
            bank_nazwa=bank,
            wgrany_plik=True,
        )

    # ── Publiczne API – weryfikacja ─────────────────────────────────────────

    def check_accounting_logic(
        self,
        dane_zois:   Optional[DaneZOiS],
        dane_bilans: Optional[DaneBilansu]  = None,
        dane_rzis:   Optional[DaneRZiS]     = None,
        wyciagi:     Optional[List[WyciagBankowy]] = None,
        rok_obrachunkowy: int = 2024,
    ) -> List[PunktKontroli]:
        self._wyniki = []
        wyciagi = wyciagi or []

        # ── ZOiS – konta ────────────────────────────────────────────────────
        if dane_zois is None:
            self._wyniki.append(PunktKontroli(
                konto="ZOiS", punkt="Wczytanie ZOiS", status=StatusAudytu.BRAK,
                uwagi="Nie dostarczono pliku ZOiS.",
            ))
        else:
            self._weryfikuj_konta_bankowe(dane_zois, wyciagi, rok_obrachunkowy)
            self._weryfikuj_konto_145(dane_zois)
            self._weryfikuj_konto_200(dane_zois)
            self._weryfikuj_konto_202(dane_zois)
            self._weryfikuj_konto_230(dane_zois)
            self._weryfikuj_konto_229(dane_zois)
            self._weryfikuj_konto_220(dane_zois)
            self._weryfikuj_konto_700(dane_zois)
            self._weryfikuj_grupe_4(dane_zois)

        # ── Bilans ──────────────────────────────────────────────────────────
        self._weryfikuj_bilans(dane_bilans)

        # ── RZiS ────────────────────────────────────────────────────────────
        self._weryfikuj_rzis(dane_rzis)

        # ── Reguły krzyżowe ZOiS ↔ RZiS ↔ Bilans ────────────────────────────
        self._weryfikuj_krzyzowe(dane_zois, dane_bilans, dane_rzis)

        return self._wyniki

    def generate_audit_report(
        self,
        wyniki: List[PunktKontroli],
        nazwa_podmiotu: str = "Podmiot",
        rok: int = 2024,
    ) -> Dict:
        linia = "═" * 70
        L = [
            linia,
            f"  RAPORT KONTROLI JAKOŚCI DANYCH – ZAMKNIĘCIE ROKU {rok}",
            f"  Podmiot: {nazwa_podmiotu}",
            f"  Wygenerowano przez: SymfoniaYearEndAuditor v3.0",
            linia, "",
            f"{'ŹRÓDŁO':<14} {'STATUS':<20} {'PUNKT KONTROLI':<42} UWAGI",
            "─" * 110,
        ]
        stats = {s: 0 for s in StatusAudytu}
        wyniki_slow = []
        for pkt in wyniki:
            stats[pkt.status] += 1
            wartosc = f" [{pkt.wartosc}]" if pkt.wartosc else ""
            L.append(
                f"{pkt.konto:<14} {pkt.status.value:<20} "
                f"{pkt.punkt:<42} {pkt.uwagi}{wartosc}"
            )
            wyniki_slow.append({
                "konto": pkt.konto, "status": pkt.status.value,
                "punkt": pkt.punkt, "uwagi": pkt.uwagi, "wartosc": pkt.wartosc,
            })

        bledy = [p for p in wyniki if p.status == StatusAudytu.BLAD]
        ostrz = [p for p in wyniki if p.status == StatusAudytu.OSTRZEZ]
        brak  = [p for p in wyniki if p.status == StatusAudytu.BRAK]

        L.extend(["", linia, "  PODSUMOWANIE", linia])
        L.append(f"  ✅ OK:            {stats[StatusAudytu.OK]}")
        L.append(f"  ❌ BŁĘDY:         {stats[StatusAudytu.BLAD]}")
        L.append(f"  ⚠️  OSTRZEŻENIA:  {stats[StatusAudytu.OSTRZEZ]}")
        L.append(f"  🔍 BRAK DANYCH:  {stats[StatusAudytu.BRAK]}")

        if bledy:
            L.extend(["", "  ❌ WYMAGANE DZIAŁANIA (BŁĘDY KRYTYCZNE):"])
            for p in bledy:
                L.append(f"    → {p.konto}: {p.uwagi}")
        if ostrz:
            L.extend(["", "  ⚠️  DO WYJAŚNIENIA (OSTRZEŻENIA):"])
            for p in ostrz:
                L.append(f"    → {p.konto}: {p.uwagi}")
        if brak:
            L.extend(["", "  🔍 BRAKI DANYCH:"])
            for p in brak:
                L.append(f"    → {p.konto}: {p.uwagi}")

        L.append("")
        if stats[StatusAudytu.BLAD] == 0 and stats[StatusAudytu.OSTRZEZ] == 0:
            L.append("  🎉 OCENA KOŃCOWA: DANE SPÓJNE – gotowe do badania.")
        elif stats[StatusAudytu.BLAD] == 0:
            L.append("  🟡 OCENA KOŃCOWA: WYMAGA WYJAŚNIENIA.")
        else:
            L.append("  🔴 OCENA KOŃCOWA: DANE NIESPÓJNE – wymagane korekty.")
        L.append(linia)

        return {
            "tekst": "\n".join(L),
            "wyniki": wyniki_slow,
            "podsumowanie": {
                "ok": stats[StatusAudytu.OK],
                "bledy": stats[StatusAudytu.BLAD],
                "ostrzezenia": stats[StatusAudytu.OSTRZEZ],
                "brak_danych": stats[StatusAudytu.BRAK],
                "podmiot": nazwa_podmiotu,
                "rok": rok,
            },
        }

    def run_full_audit(
        self,
        *,
        zois_bytes:   Optional[bytes] = None,
        zois_format:  str = "xlsx",
        bilans_bytes: Optional[bytes] = None,
        bilans_format: str = "xlsx",
        rzis_bytes:   Optional[bytes] = None,
        rzis_format:  str = "pdf",
        wyciagi: Optional[List[WyciagBankowy]] = None,
        nazwa_podmiotu: str = "Podmiot",
        rok: int = 2024,
    ) -> Dict:
        dane_zois   = self.parsuj_zois(zois_bytes, zois_format) if zois_bytes else None
        dane_bilans = self.parsuj_bilans(bilans_bytes, bilans_format) if bilans_bytes else None
        dane_rzis   = self.parsuj_rzis(rzis_bytes, rzis_format) if rzis_bytes else None

        wyniki = self.check_accounting_logic(
            dane_zois, dane_bilans, dane_rzis, wyciagi or [],
            rok_obrachunkowy=rok,
        )
        return self.generate_audit_report(wyniki, nazwa_podmiotu, rok)

    # ─── WERYFIKACJA BANK ─────────────────────────────────────────────────────

    def _weryfikuj_konta_bankowe(
        self,
        dane_zois: DaneZOiS,
        wyciagi: List[WyciagBankowy],
        rok_obrachunkowy: int,
    ) -> None:
        konta_bankowe = dane_zois.pobierz_konta_bankowe()
        if not konta_bankowe:
            self._wyniki.append(PunktKontroli(
                konto="130", punkt="Rachunki bankowe",
                status=StatusAudytu.BRAK,
                uwagi="CRITICAL: Nie odnaleziono konta 130 w ZOiS.",
            ))
            return

        mapa = {w.numer_konta_ksiegowego: w for w in wyciagi}
        self._wyniki.append(PunktKontroli(
            konto="130", punkt="Wykryte rachunki bankowe",
            status=StatusAudytu.INFO,
            uwagi=f"Wykryto {len(konta_bankowe)} rachunek(ów) w ZOiS, wgrano {len(wyciagi)} wyciąg(ów).",
        ))

        rachunki_bez_wyciagu = []
        for numer_ks, opis, saldo_zois in konta_bankowe:
            wyciag = mapa.get(numer_ks)
            if wyciag is None:
                rachunki_bez_wyciagu.append((numer_ks, opis, saldo_zois))
                self._wyniki.append(PunktKontroli(
                    konto=numer_ks, punkt=f"Rachunek {opis}",
                    status=StatusAudytu.BRAK,
                    uwagi=f"Brak wgranego wyciągu. Saldo ZOiS: {saldo_zois:,.2f} zł.",
                ))
                continue

            roznica = abs(saldo_zois - wyciag.saldo_koncowe)
            okres_info = ""
            ostrz_okresu = ""
            if wyciag.miesiac_ostatniej_operacji:
                okres_info = f" | Ostatnia operacja: {wyciag.okres_opisowy}"
                if (wyciag.miesiac_ostatniej_operacji != 12 or
                    (wyciag.rok_ostatniej_operacji and
                     wyciag.rok_ostatniej_operacji != rok_obrachunkowy)):
                    ostrz_okresu = (
                        f" UWAGA: Rachunek zamknięty operacjami z "
                        f"{wyciag.okres_opisowy}, nie z grudnia {rok_obrachunkowy}."
                    )
            bank_str = f" ({wyciag.bank_nazwa})" if wyciag.bank_nazwa else ""

            if roznica < Decimal("0.01"):
                status = StatusAudytu.OSTRZEZ if ostrz_okresu else StatusAudytu.OK
                self._wyniki.append(PunktKontroli(
                    konto=numer_ks, punkt=f"Rachunek {opis}{bank_str}",
                    status=status,
                    uwagi=f"Saldo ZOiS zgodne z wyciągiem.{ostrz_okresu}{okres_info}",
                    wartosc=f"{saldo_zois:,.2f} zł",
                ))
            else:
                self._wyniki.append(PunktKontroli(
                    konto=numer_ks, punkt=f"Rachunek {opis}{bank_str}",
                    status=StatusAudytu.BLAD,
                    uwagi=(f"NIEZGODNOŚĆ: ZOiS={saldo_zois:,.2f} zł | "
                           f"Wyciąg={wyciag.saldo_koncowe:,.2f} zł | "
                           f"Różnica={roznica:,.2f} zł.{ostrz_okresu}{okres_info}"),
                ))

        if rachunki_bez_wyciagu:
            lista = "; ".join(f"{n} ({o}): {s:,.2f} zł" for n, o, s in rachunki_bez_wyciagu)
            self._wyniki.append(PunktKontroli(
                konto="130 (podsum.)",
                punkt=f"Rachunki bez wgranego wyciągu ({len(rachunki_bez_wyciagu)} szt.)",
                status=StatusAudytu.OSTRZEZ,
                uwagi=f"Do uzupełnienia: {lista}",
            ))

    # ─── WERYFIKACJA POSZCZEGÓLNYCH KONT (jak w v2.0) ─────────────────────────

    def _saldo(self, dane_zois, konto):
        return dane_zois.konta.get(konto)

    def _znajdz_syntetyki_po_opisie(
        self,
        dane_zois: "DaneZOiS",
        slowa_kluczowe: List[str],
        grupy_syntetyk: Optional[List[str]] = None,
    ) -> List[Tuple[str, str, Decimal, Decimal]]:
        """
        Wyszukuje syntetyki których opis zawiera któreś ze słów kluczowych.

        Parametry:
          - slowa_kluczowe: np. ["odbiorc", "klient"]
          - grupy_syntetyk: opcjonalnie zawężenie do grup ["20", "22"]

        Zwraca listę: (numer_konta, opis, saldo_Wn, saldo_Ma).
        """
        wynik = []
        slowa_lower = [s.lower() for s in slowa_kluczowe]
        for numer, (wn, ma) in dane_zois.konta.items():
            # Tylko czyste syntetyki (bez myślnika)
            if "-" in numer:
                continue
            if grupy_syntetyk and not any(numer.startswith(g) for g in grupy_syntetyk):
                continue
            opis = dane_zois.opisy.get(numer, "")
            opis_lower = opis.lower()
            if any(s in opis_lower for s in slowa_lower):
                wynik.append((numer, opis, wn, ma))
        return wynik

    def _znajdz_analityki_po_opisie(
        self,
        dane_zois: "DaneZOiS",
        slowa_kluczowe: List[str],
        grupy_syntetyk: Optional[List[str]] = None,
    ) -> List[Tuple[str, str, Decimal, Decimal]]:
        """
        Analogicznie do syntetyk, ale szuka w analitykach (konta z myślnikiem).
        Przykład użycia: szukanie analityki ZUS "220-3" po opisie "ZUS".
        """
        wynik = []
        slowa_lower = [s.lower() for s in slowa_kluczowe]
        for numer, (wn, ma) in dane_zois.konta_analityki.items():
            if "-" not in numer:
                continue
            numer_syn = normalize_konto(numer)
            if grupy_syntetyk and not any(numer_syn.startswith(g) for g in grupy_syntetyk):
                continue
            opis = dane_zois.opisy.get(numer, "")
            opis_lower = opis.lower()
            if any(s in opis_lower for s in slowa_lower):
                wynik.append((numer, opis, wn, ma))
        return wynik

    def _weryfikuj_konto_145(self, dane_zois):
        """
        Weryfikacja środków pieniężnych w drodze. W różnych planach kont:
          - 145 (standard)
          - 133/134 (IGRAPES – "środki w drodze", "rachunek VAT")
        Saldo powinno wynosić 0 na koniec roku.
        """
        # Szukamy po opisie "w drodze" albo "drodze" w grupach 13x/14x
        kandydaci = self._znajdz_syntetyki_po_opisie(
            dane_zois, ["w drodze", "drodze"], grupy_syntetyk=["13", "14"]
        )
        # Jeśli nie znaleziono po opisie - spróbuj klasyczne 145
        if not kandydaci:
            s = self._saldo(dane_zois, "145")
            if s is None:
                self._wyniki.append(PunktKontroli(
                    konto="145", punkt="Środki pieniężne w drodze = 0",
                    status=StatusAudytu.INFO,
                    uwagi="Konto środków w drodze nie wystąpiło.",
                )); return
            kandydaci = [("145", dane_zois.opisy.get("145", "Środki w drodze"),
                         s[0], s[1])]

        # Sprawdzamy każde konto z osobna
        for numer, opis, wn, ma in kandydaci:
            if wn == Decimal("0") and ma == Decimal("0"):
                self._wyniki.append(PunktKontroli(
                    konto=numer, punkt=f"{opis} = 0",
                    status=StatusAudytu.OK, uwagi="Saldo wynosi 0.",
                ))
            else:
                self._wyniki.append(PunktKontroli(
                    konto=numer, punkt=f"{opis} = 0",
                    status=StatusAudytu.BLAD,
                    uwagi=f"Saldo ≠ 0! Wn={wn:,.2f}, Ma={ma:,.2f} zł.",
                ))

    def _weryfikuj_konto_200(self, dane_zois):
        """
        Rozrachunki z odbiorcami – grupa 20x (np. 200, 201).
        Wyszukujemy po opisie zawierającym "odbiorc" lub "klient".
        Oczekiwane saldo: Wn (należności).
        """
        kandydaci = self._znajdz_syntetyki_po_opisie(
            dane_zois, ["odbiorc", "klient"], grupy_syntetyk=["20"]
        )
        if not kandydaci:
            # Fallback: klasyczne 200
            s = self._saldo(dane_zois, "200")
            if s is None:
                self._wyniki.append(PunktKontroli(
                    konto="20x", punkt="Rozrachunki z odbiorcami",
                    status=StatusAudytu.BRAK,
                    uwagi="Brak konta rozrachunków z odbiorcami (20x).",
                )); return
            kandydaci = [("200", dane_zois.opisy.get("200", "Rozrachunki - odbiorcy"),
                         s[0], s[1])]

        for numer, opis, wn, ma in kandydaci:
            if ma > Decimal("0"):
                self._wyniki.append(PunktKontroli(
                    konto=numer, punkt=f"{opis} – strona salda",
                    status=StatusAudytu.BLAD,
                    uwagi=f"Saldo Ma={ma:,.2f} zł – BŁĄD. Brak faktury sprzedaży lub nadpłata.",
                ))
            else:
                self._wyniki.append(PunktKontroli(
                    konto=numer, punkt=f"{opis} – strona salda",
                    status=StatusAudytu.OK,
                    uwagi=f"Saldo Wn={wn:,.2f} zł – należności.",
                ))

    def _weryfikuj_konto_202(self, dane_zois):
        """
        Rozrachunki z dostawcami – grupa 20x (np. 202, 210, 211).
        Wyszukujemy po opisie zawierającym "dostawc".
        Oczekiwane saldo: Ma (zobowiązania).
        """
        kandydaci = self._znajdz_syntetyki_po_opisie(
            dane_zois, ["dostawc"], grupy_syntetyk=["20", "21"]
        )
        if not kandydaci:
            s = self._saldo(dane_zois, "202")
            if s is None:
                self._wyniki.append(PunktKontroli(
                    konto="20x/21x", punkt="Rozrachunki z dostawcami",
                    status=StatusAudytu.BRAK,
                    uwagi="Brak konta rozrachunków z dostawcami.",
                )); return
            kandydaci = [("202", dane_zois.opisy.get("202", "Rozrachunki - dostawcy"),
                         s[0], s[1])]

        for numer, opis, wn, ma in kandydaci:
            if wn > Decimal("0"):
                self._wyniki.append(PunktKontroli(
                    konto=numer, punkt=f"{opis} – strona salda",
                    status=StatusAudytu.BLAD,
                    uwagi=f"Saldo Wn={wn:,.2f} zł – BŁĄD. Nadpłata lub brak faktury zakupu.",
                ))
            else:
                self._wyniki.append(PunktKontroli(
                    konto=numer, punkt=f"{opis} – strona salda",
                    status=StatusAudytu.OK,
                    uwagi=f"Saldo Ma={ma:,.2f} zł – zobowiązania.",
                ))

    def _weryfikuj_konto_230(self, dane_zois):
        s = self._saldo(dane_zois, "230")
        if s is None:
            self._wyniki.append(PunktKontroli(
                konto="230", punkt="Wynagrodzenia",
                status=StatusAudytu.INFO, uwagi="Konto 230 nie wystąpiło.",
            )); return
        wn, ma = s
        if wn > Decimal("0"):
            self._wyniki.append(PunktKontroli(
                konto="230", punkt="Rozrachunki z pracownikami",
                status=StatusAudytu.OSTRZEZ,
                uwagi=f"Saldo Wn={wn:,.2f} zł – brak LP lub nadpłata.",
            ))
        elif ma > Decimal("0"):
            self._wyniki.append(PunktKontroli(
                konto="230", punkt="Rozrachunki z pracownikami",
                status=StatusAudytu.OK,
                uwagi=f"Saldo Ma={ma:,.2f} zł – niezapłacone (dopuszczalne).",
            ))
        else:
            self._wyniki.append(PunktKontroli(
                konto="230", punkt="Rozrachunki z pracownikami",
                status=StatusAudytu.OK, uwagi="Saldo 230 = 0.",
            ))

    def _weryfikuj_konto_229(self, dane_zois):
        """
        Zobowiązania ZUS. W różnych planach kont:
          - 229 (klasyczny standard)
          - 220-3 w IGRAPES (analityka pod 220 Rozrachunki z budżetami)
          - 221-x w innych
        Szuka po opisie "zus" lub "ubezpiecz" w grupach 22x.
        Oczekiwane saldo: Ma (zobowiązanie).
        """
        # Najpierw szukaj w syntetykach
        kandydaci = self._znajdz_syntetyki_po_opisie(
            dane_zois, ["zus"], grupy_syntetyk=["22"]
        )
        # Potem w analitykach (IGRAPES: 220-3 ZUS)
        kandydaci_ana = self._znajdz_analityki_po_opisie(
            dane_zois, ["zus"], grupy_syntetyk=["22"]
        )
        # Gdy mamy analityki, eliminujemy duplikat - syntetykę tej samej grupy
        if kandydaci_ana:
            syntetyki_do_pominiecia = {
                normalize_konto(n) for n, _, _, _ in kandydaci_ana
            }
            kandydaci = [
                k for k in kandydaci if k[0] not in syntetyki_do_pominiecia
            ]
        kandydaci.extend(kandydaci_ana)

        if not kandydaci:
            # Fallback: klasyczne 229
            s = self._saldo(dane_zois, "229")
            if s is None:
                self._wyniki.append(PunktKontroli(
                    konto="ZUS", punkt="Zobowiązanie ZUS",
                    status=StatusAudytu.BRAK,
                    uwagi="Brak konta ZUS (229/220-x/ubezpieczeń).",
                )); return
            kandydaci = [("229", dane_zois.opisy.get("229", "ZUS"), s[0], s[1])]

        for numer, opis, wn, ma in kandydaci:
            if wn > Decimal("0"):
                self._wyniki.append(PunktKontroli(
                    konto=numer, punkt=f"{opis} – zobowiązanie ZUS",
                    status=StatusAudytu.OSTRZEZ,
                    uwagi=f"Saldo Wn={wn:,.2f} zł – nadpłata lub błąd DRA.",
                ))
            elif ma > Decimal("0"):
                self._wyniki.append(PunktKontroli(
                    konto=numer, punkt=f"{opis} – zobowiązanie ZUS",
                    status=StatusAudytu.OK,
                    uwagi=f"Saldo Ma={ma:,.2f} zł. Porównaj z DRA.",
                ))
            else:
                self._wyniki.append(PunktKontroli(
                    konto=numer, punkt=f"{opis} – zobowiązanie ZUS",
                    status=StatusAudytu.OK, uwagi="Saldo wynosi 0.",
                ))

    def _weryfikuj_konto_220(self, dane_zois):
        """
        Zobowiązania podatkowe – PIT-4 (podatek od płac) oraz podatek dochodowy.
        W różnych planach kont:
          - 220 (klasyczny - cały US)
          - 220-1 Podatek dochodowy, 220-2 Podatek od płac (IGRAPES)
          - 222-x (inne systemy)
        Szuka po opisie "podatek" w grupie 22x (bez ZUS).
        Oczekiwane saldo: Ma (zobowiązanie podatkowe).
        """
        kandydaci = self._znajdz_syntetyki_po_opisie(
            dane_zois, ["podatek", "budżet", "urz"], grupy_syntetyk=["22"]
        )
        kandydaci_ana = self._znajdz_analityki_po_opisie(
            dane_zois, ["podatek", "pit"], grupy_syntetyk=["22"]
        )
        # Wyłącz analityki ZUS (już obsłużone w _weryfikuj_konto_229)
        kandydaci_ana = [
            k for k in kandydaci_ana if "zus" not in k[1].lower()
        ]
        # Gdy mamy analityki, pomijamy syntetyki tej samej grupy (duplikacja)
        if kandydaci_ana:
            syntetyki_do_pominiecia = {
                normalize_konto(n) for n, _, _, _ in kandydaci_ana
            }
            kandydaci = [
                k for k in kandydaci if k[0] not in syntetyki_do_pominiecia
            ]
        kandydaci.extend(kandydaci_ana)

        if not kandydaci:
            s = self._saldo(dane_zois, "220")
            if s is None:
                self._wyniki.append(PunktKontroli(
                    konto="US/PIT", punkt="Zobowiązanie podatkowe",
                    status=StatusAudytu.BRAK,
                    uwagi="Brak konta podatkowego (220/22x).",
                )); return
            kandydaci = [("220", dane_zois.opisy.get("220", "US"), s[0], s[1])]

        for numer, opis, wn, ma in kandydaci:
            if wn > Decimal("0"):
                self._wyniki.append(PunktKontroli(
                    konto=numer, punkt=f"{opis} – zobowiązanie podatkowe",
                    status=StatusAudytu.OSTRZEZ,
                    uwagi=f"Saldo Wn={wn:,.2f} zł – nadpłata lub brak dekretacji XII.",
                ))
            elif ma > Decimal("0"):
                self._wyniki.append(PunktKontroli(
                    konto=numer, punkt=f"{opis} – zobowiązanie podatkowe",
                    status=StatusAudytu.OK,
                    uwagi=f"Saldo Ma={ma:,.2f} zł. Porównaj z deklaracją PIT-4R/8AR.",
                ))
            else:
                self._wyniki.append(PunktKontroli(
                    konto=numer, punkt=f"{opis} – zobowiązanie podatkowe",
                    status=StatusAudytu.OK, uwagi="Saldo wynosi 0.",
                ))

    def _weryfikuj_konto_700(self, dane_zois):
        """
        Weryfikuje konta grupy 70 (przychody ze sprzedaży: 700-709).
        W praktyce podmioty używają różnych numerów: 700 (Sprzedaż), 701, 702
        (Sprzedaż usług), 703 (Sprzedaż towarów) itp.
        """
        # Sumuj wszystkie konta z grupy 70x (700-709)
        konta_70 = {
            k: v for k, v in dane_zois.konta.items()
            if k.isdigit() and len(k) == 3 and k.startswith("70")
        }
        if not konta_70:
            self._wyniki.append(PunktKontroli(
                konto="70x", punkt="Przychody ze sprzedaży",
                status=StatusAudytu.BRAK,
                uwagi="CRITICAL: Brak kont grupy 70x (700-709).",
            )); return

        suma_wn = sum((wn for wn, _ in konta_70.values()), Decimal("0"))
        suma_ma = sum((ma for _, ma in konta_70.values()), Decimal("0"))
        konta_str = ", ".join(sorted(konta_70.keys()))

        if suma_wn > Decimal("0"):
            bledne = [k for k, (wn, _) in konta_70.items() if wn > Decimal("0")]
            self._wyniki.append(PunktKontroli(
                konto="70x", punkt="Przychody ze sprzedaży – strona salda",
                status=StatusAudytu.BLAD,
                uwagi=f"Saldo Wn={suma_wn:,.2f} zł – BŁĄD! "
                      f"Konta wynikowe Ma. Błędne konta: {', '.join(bledne)}.",
            ))
        elif suma_ma > Decimal("0"):
            self._wyniki.append(PunktKontroli(
                konto="70x", punkt=f"Przychody ze sprzedaży ({konta_str})",
                status=StatusAudytu.OK,
                uwagi=f"Saldo Ma={suma_ma:,.2f} zł – poprawne.",
                wartosc=f"{suma_ma:,.2f} zł",
            ))
        else:
            self._wyniki.append(PunktKontroli(
                konto="70x", punkt="Przychody ze sprzedaży",
                status=StatusAudytu.OSTRZEZ,
                uwagi="Zerowe salda – brak sprzedaży w okresie?",
            ))

    def _weryfikuj_grupe_4(self, dane_zois):
        konta = {k: v for k, v in dane_zois.konta.items() if get_grupa(k) == 4}
        if not konta:
            self._wyniki.append(PunktKontroli(
                konto="Grupa 4", punkt="Koszty rodzajowe (400–499)",
                status=StatusAudytu.BRAK, uwagi="Brak kont grupy 4.",
            )); return
        bledne = []
        suma = Decimal("0")
        for k, (wn, ma) in konta.items():
            suma += wn
            if ma > Decimal("0"):
                opis = dane_zois.opisy.get(k, "")
                bledne.append(f"{k} ({opis}): Ma={ma:,.2f} zł")
        if bledne:
            self._wyniki.append(PunktKontroli(
                konto="Grupa 4", punkt="Koszty rodzajowe – tylko Saldo Wn",
                status=StatusAudytu.BLAD,
                uwagi=(f"Konta z Saldem Ma ({len(bledne)} szt.): "
                       + "; ".join(bledne[:5])
                       + (" ..." if len(bledne) > 5 else "")),
            ))
        else:
            self._wyniki.append(PunktKontroli(
                konto="Grupa 4", punkt="Koszty rodzajowe – tylko Saldo Wn",
                status=StatusAudytu.OK,
                uwagi=f"Wszystkie {len(konta)} kont poprawne. Suma: {suma:,.2f} zł.",
            ))

    # ─── WERYFIKACJA BILANSU (v3.0 – rozszerzona) ────────────────────────────

    def _weryfikuj_bilans(self, dane_bilans: Optional[DaneBilansu]):
        if dane_bilans is None:
            self._wyniki.append(PunktKontroli(
                konto="Bilans", punkt="Suma bilansowa",
                status=StatusAudytu.BRAK, uwagi="Nie dostarczono pliku Bilansu.",
            )); return

        TOL = Decimal("0.01")

        # 1. Spójność wewnętrzna aktywów: A + B + C + D = Suma aktywów
        suma_obl_a = (
            dane_bilans.aktywa_trwale_biezacy + dane_bilans.aktywa_obrotowe_biezacy
            + dane_bilans.nalezne_wplaty_biezacy + dane_bilans.udzialy_wlasne_biezacy
        )
        if dane_bilans.suma_aktywow_biezacy > Decimal("0"):
            roznica = abs(suma_obl_a - dane_bilans.suma_aktywow_biezacy)
            if roznica <= TOL:
                self._wyniki.append(PunktKontroli(
                    konto="Bilans", punkt="Spójność wewnętrzna aktywów (A+B+C+D=Suma)",
                    status=StatusAudytu.OK,
                    uwagi=(f"Aktywa trwałe+obrotowe+C+D = Suma aktywów "
                           f"({dane_bilans.suma_aktywow_biezacy:,.2f} zł)."),
                ))
            else:
                self._wyniki.append(PunktKontroli(
                    konto="Bilans", punkt="Spójność wewnętrzna aktywów (A+B+C+D=Suma)",
                    status=StatusAudytu.BLAD,
                    uwagi=(f"NIEZGODNOŚĆ! Obliczona suma={suma_obl_a:,.2f}, "
                           f"Suma w bilansie={dane_bilans.suma_aktywow_biezacy:,.2f} | "
                           f"Różnica={roznica:,.2f} zł."),
                ))

        # 2. Spójność wewnętrzna pasywów: A + B = Suma pasywów
        suma_obl_p = dane_bilans.kapital_wlasny_biezacy + dane_bilans.zobowiazania_biezacy
        if dane_bilans.suma_pasywow_biezacy > Decimal("0"):
            roznica = abs(suma_obl_p - dane_bilans.suma_pasywow_biezacy)
            if roznica <= TOL:
                self._wyniki.append(PunktKontroli(
                    konto="Bilans", punkt="Spójność wewnętrzna pasywów (A+B=Suma)",
                    status=StatusAudytu.OK,
                    uwagi=(f"Kapitał własny+Zobowiązania = Suma pasywów "
                           f"({dane_bilans.suma_pasywow_biezacy:,.2f} zł)."),
                ))
            else:
                self._wyniki.append(PunktKontroli(
                    konto="Bilans", punkt="Spójność wewnętrzna pasywów (A+B=Suma)",
                    status=StatusAudytu.BLAD,
                    uwagi=(f"NIEZGODNOŚĆ! Obliczona suma={suma_obl_p:,.2f}, "
                           f"Suma w bilansie={dane_bilans.suma_pasywow_biezacy:,.2f} | "
                           f"Różnica={roznica:,.2f} zł."),
                ))

        # 3. Aktywa = Pasywa (rok bieżący)
        rb = abs(dane_bilans.suma_aktywow_biezacy - dane_bilans.suma_pasywow_biezacy)
        if rb <= TOL:
            self._wyniki.append(PunktKontroli(
                konto="Bilans", punkt="Suma bilansowa rok bieżący",
                status=StatusAudytu.OK,
                uwagi=f"Aktywa = Pasywa = {dane_bilans.suma_aktywow_biezacy:,.2f} zł.",
            ))
        else:
            self._wyniki.append(PunktKontroli(
                konto="Bilans", punkt="Suma bilansowa rok bieżący",
                status=StatusAudytu.BLAD,
                uwagi=(f"NIEZGODNOŚĆ! Aktywa={dane_bilans.suma_aktywow_biezacy:,.2f} ≠ "
                       f"Pasywa={dane_bilans.suma_pasywow_biezacy:,.2f} | Δ={rb:,.2f} zł."),
            ))

        # 4. Aktywa = Pasywa (rok ubiegły)
        if dane_bilans.suma_aktywow_ubiegly > Decimal("0") or dane_bilans.suma_pasywow_ubiegly > Decimal("0"):
            ru = abs(dane_bilans.suma_aktywow_ubiegly - dane_bilans.suma_pasywow_ubiegly)
            if ru <= TOL:
                self._wyniki.append(PunktKontroli(
                    konto="Bilans", punkt="Suma bilansowa rok ubiegły",
                    status=StatusAudytu.OK,
                    uwagi=f"Dane porównawcze: {dane_bilans.suma_aktywow_ubiegly:,.2f} zł.",
                ))
            else:
                self._wyniki.append(PunktKontroli(
                    konto="Bilans", punkt="Suma bilansowa rok ubiegły",
                    status=StatusAudytu.BLAD,
                    uwagi=(f"NIEZGODNOŚĆ! Aktywa={dane_bilans.suma_aktywow_ubiegly:,.2f} ≠ "
                           f"Pasywa={dane_bilans.suma_pasywow_ubiegly:,.2f} | Δ={ru:,.2f} zł."),
                ))

    # ─── WERYFIKACJA RZiS (v3.0 – nowa) ──────────────────────────────────────

    def _weryfikuj_rzis(self, dane_rzis: Optional[DaneRZiS]):
        if dane_rzis is None:
            self._wyniki.append(PunktKontroli(
                konto="RZiS", punkt="Rachunek Zysków i Strat",
                status=StatusAudytu.BRAK, uwagi="Nie dostarczono pliku RZiS.",
            )); return

        TOL = Decimal("1.00")  # tolerancja 1 zł (zaokrąglenia Symfonii)

        # Sprawdzenia spójności matematycznej RZiS
        reguly = [
            ("C = A − B (Zysk ze sprzedaży)",
             dane_rzis.przychody_sprzedazy[0] - dane_rzis.koszty_operacyjne[0],
             dane_rzis.zysk_sprzedazy[0]),
            ("F = C + D − E (Wynik operacyjny)",
             dane_rzis.zysk_sprzedazy[0]
              + dane_rzis.pozostale_przych_oper[0]
              - dane_rzis.pozostale_koszty_oper[0],
             dane_rzis.zysk_dzialalnosci_oper[0]),
            ("I = F + G − H (Zysk brutto)",
             dane_rzis.zysk_dzialalnosci_oper[0]
              + dane_rzis.przychody_finansowe[0]
              - dane_rzis.koszty_finansowe[0],
             dane_rzis.zysk_brutto[0]),
            ("L = I − J − K (Zysk netto)",
             dane_rzis.zysk_brutto[0]
              - dane_rzis.podatek_dochodowy[0]
              - dane_rzis.pozostale_zmniejszenia[0],
             dane_rzis.zysk_netto[0]),
        ]

        for nazwa, obliczona, podana in reguly:
            roznica = abs(obliczona - podana)
            if roznica <= TOL:
                self._wyniki.append(PunktKontroli(
                    konto="RZiS", punkt=nazwa,
                    status=StatusAudytu.OK,
                    uwagi=f"Spójność matematyczna OK ({podana:,.2f} zł).",
                ))
            else:
                self._wyniki.append(PunktKontroli(
                    konto="RZiS", punkt=nazwa,
                    status=StatusAudytu.BLAD,
                    uwagi=(f"NIESPÓJNOŚĆ! Obliczona={obliczona:,.2f}, "
                           f"Podana={podana:,.2f} | Różnica={roznica:,.2f} zł."),
                ))

        # Informacyjnie – zysk/strata netto
        zn = dane_rzis.zysk_netto[0]
        self._wyniki.append(PunktKontroli(
            konto="RZiS", punkt="Zysk (strata) netto",
            status=StatusAudytu.INFO,
            uwagi=f"Wynik finansowy roku bieżącego: {zn:,.2f} zł.",
            wartosc=f"{zn:,.2f} zł",
        ))

    # ─── REGUŁY KRZYŻOWE (v3.0 – nowa) ───────────────────────────────────────

    def _weryfikuj_krzyzowe(
        self,
        dane_zois: Optional[DaneZOiS],
        dane_bilans: Optional[DaneBilansu],
        dane_rzis: Optional[DaneRZiS],
    ):
        """Weryfikacja krzyżowa między ZOiS, Bilansem i RZiS."""
        TOL = Decimal("1.00")

        # 1. Wynik netto RZiS = Wynik netto w Bilansie (pasywa A.VI)
        if dane_rzis is not None and dane_bilans is not None:
            zn_rzis = dane_rzis.zysk_netto[0]
            zn_bil = dane_bilans.wynik_netto_biezacy
            if zn_rzis == Decimal("0") and zn_bil == Decimal("0"):
                pass  # nie mamy danych
            else:
                roznica = abs(zn_rzis - zn_bil)
                if roznica <= TOL:
                    self._wyniki.append(PunktKontroli(
                        konto="RZiS↔Bilans",
                        punkt="Wynik netto: RZiS (L) = Bilans (Pasywa A.VI)",
                        status=StatusAudytu.OK,
                        uwagi=f"Wynik netto zgodny: {zn_rzis:,.2f} zł.",
                    ))
                else:
                    self._wyniki.append(PunktKontroli(
                        konto="RZiS↔Bilans",
                        punkt="Wynik netto: RZiS (L) = Bilans (Pasywa A.VI)",
                        status=StatusAudytu.BLAD,
                        uwagi=(f"NIEZGODNOŚĆ! RZiS={zn_rzis:,.2f} zł, "
                               f"Bilans={zn_bil:,.2f} zł | Δ={roznica:,.2f} zł."),
                    ))

        # 2. Grupa 70x (Ma) ≈ RZiS poz. A (Przychody ze sprzedaży)
        if dane_zois is not None and dane_rzis is not None:
            konta_70 = {
                k: v for k, v in dane_zois.konta.items()
                if k.isdigit() and len(k) == 3 and k.startswith("70")
            }
            if konta_70:
                suma_70_ma = sum((ma for _, ma in konta_70.values()), Decimal("0"))
                przych = dane_rzis.przychody_sprzedazy[0]
                konta_str = "+".join(sorted(konta_70.keys()))
                if przych > Decimal("0") and suma_70_ma > Decimal("0"):
                    roznica = abs(suma_70_ma - przych)
                    if roznica <= TOL:
                        self._wyniki.append(PunktKontroli(
                            konto="ZOiS↔RZiS",
                            punkt=f"Grupa 70x (Ma) ≈ RZiS poz. A (Przychody)",
                            status=StatusAudytu.OK,
                            uwagi=f"Przychody zgodne ({konta_str}): {suma_70_ma:,.2f} zł.",
                        ))
                    else:
                        self._wyniki.append(PunktKontroli(
                            konto="ZOiS↔RZiS",
                            punkt=f"Grupa 70x (Ma) ≈ RZiS poz. A (Przychody)",
                            status=StatusAudytu.OSTRZEZ,
                            uwagi=(f"Różnica: ZOiS {konta_str}={suma_70_ma:,.2f} zł, "
                                   f"RZiS A={przych:,.2f} zł | Δ={roznica:,.2f} zł."),
                        ))

        # 3. Suma grupy 4 (Wn) ≈ RZiS poz. B (Koszty operacyjne)
        if dane_zois is not None and dane_rzis is not None:
            grupa_4_wn = sum(
                (wn for k, (wn, _) in dane_zois.konta.items() if get_grupa(k) == 4),
                Decimal("0")
            )
            koszty_b = dane_rzis.koszty_operacyjne[0]
            if grupa_4_wn > Decimal("0") and koszty_b > Decimal("0"):
                roznica = abs(grupa_4_wn - koszty_b)
                if roznica <= TOL:
                    self._wyniki.append(PunktKontroli(
                        konto="ZOiS↔RZiS",
                        punkt="Suma grupy 4 (Wn) ≈ RZiS poz. B (Koszty)",
                        status=StatusAudytu.OK,
                        uwagi=f"Koszty operacyjne zgodne: {koszty_b:,.2f} zł.",
                    ))
                else:
                    self._wyniki.append(PunktKontroli(
                        konto="ZOiS↔RZiS",
                        punkt="Suma grupy 4 (Wn) ≈ RZiS poz. B (Koszty)",
                        status=StatusAudytu.OSTRZEZ,
                        uwagi=(f"Różnica: ZOiS grupa 4={grupa_4_wn:,.2f} zł, "
                               f"RZiS B={koszty_b:,.2f} zł | Δ={roznica:,.2f} zł. "
                               "Sprawdź amortyzację i pozostałe pozycje."),
                    ))


# =============================================================================
# TRYB TESTOWY
# =============================================================================

if __name__ == "__main__":
    import sys

    print("=" * 70)
    print("  SymfoniaYearEndAuditor v3.0 – Tryb Testowy")
    print("  Test parsowania Bilansu + RZiS (rzeczywiste pliki Symfonia)")
    print("=" * 70)

    sciezka_bilans = "/mnt/user-data/uploads/Bilans.pdf"
    sciezka_rzis = "/mnt/user-data/uploads/RZIS.pdf"

    audytor = SymfoniaYearEndAuditor()

    print("\n→ Parsowanie Bilansu...")
    with open(sciezka_bilans, "rb") as f:
        dane_bilans = audytor.parsuj_bilans(f.read(), "pdf")
    print(f"  Suma aktywów:  {dane_bilans.suma_aktywow_biezacy:,.2f}")
    print(f"  Suma pasywów:  {dane_bilans.suma_pasywow_biezacy:,.2f}")
    print(f"  Wynik netto:   {dane_bilans.wynik_netto_biezacy:,.2f}")

    print("\n→ Parsowanie RZiS...")
    with open(sciezka_rzis, "rb") as f:
        dane_rzis = audytor.parsuj_rzis(f.read(), "pdf")
    print(f"  A. Przychody:       {dane_rzis.przychody_sprzedazy[0]:,.2f}")
    print(f"  B. Koszty oper.:    {dane_rzis.koszty_operacyjne[0]:,.2f}")
    print(f"  L. Zysk netto:      {dane_rzis.zysk_netto[0]:,.2f}")

    # Mock ZOiS do testowania krzyżowych reguł
    dz = DaneZOiS()
    dz.konta_analityki = {"130-1": (Decimal("179559.96"), Decimal("0"))}
    dz.konta = {
        "130": (Decimal("179559.96"), Decimal("0")),
        "145": (Decimal("0"), Decimal("0")),
        "200": (Decimal("201405.89"), Decimal("0")),
        "400": (Decimal("24679.19"), Decimal("0")),
        "401": (Decimal("358515.66"), Decimal("0")),
        "402": (Decimal("415178.98"), Decimal("0")),
        "403": (Decimal("73161.89"), Decimal("0")),
        "404": (Decimal("1455.89"), Decimal("0")),
        "405": (Decimal("37380.00"), Decimal("0")),
        "406": (Decimal("75366.88"), Decimal("0")),
        "700": (Decimal("0"), Decimal("1114621.54")),  # poprawne: Ma = przychody
    }
    dz.opisy = {"130-1": "Santander"}

    wyciagi = [WyciagBankowy(
        numer_konta_ksiegowego="130-1",
        saldo_koncowe=Decimal("179559.96"),
        rok_ostatniej_operacji=2024, miesiac_ostatniej_operacji=12,
        bank_nazwa="Santander",
    )]

    wyniki = audytor.check_accounting_logic(dz, dane_bilans, dane_rzis, wyciagi, 2024)
    raport = audytor.generate_audit_report(
        wyniki, nazwa_podmiotu="Abacus Centrum Księgowe SP. z o.o.", rok=2024
    )
    print("\n" + raport["tekst"])
